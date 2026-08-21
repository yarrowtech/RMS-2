
"""
vendor_routes.py
==================
⚠️ SCHEMA CHANGE — vendor identity is now separate from tenant membership.

Previously: vendors_collection had ONE document per vendor, with a single
tenant_id + status + vendor_code baked directly onto it. A vendor could
only ever belong to one retailer.

Now: vendors_collection holds only IDENTITY — name, contact info, PAN/GST,
email + password for login. A NEW collection, vendor_tenant_links_collection,
holds one document per (vendor_id, tenant_id) pair — that's where status
(Pending/Approved/Rejected), source (invite_link/self_registration/
walkin_po_self_register), vendor_code, division/section/department, and
approval timestamps now live. A vendor with relationships to both Citimart
and Zudio has ONE identity document and TWO link documents — one login,
multiple retailers, each with its own independent approval status.

REQUIRED — add to db.py:
    vendor_tenant_links_collection = db["vendor_tenant_links"]

Recommended index (uniqueness + fast lookups):
    await vendor_tenant_links_collection.create_index(
        [("vendor_id", 1), ("tenant_id", 1)], unique=True
    )

MIGRATION: existing vendor documents (from before this change) still carry
tenant_id/status/vendor_code directly on the vendors_collection doc. See
migrate_vendor_tenant_links.py — it must run once, before deploying this
file, to split every existing vendor document into an identity doc + a
link doc. Until that migration runs, existing vendors will have no link
document and will not appear in any tenant's pending/approved lists.

FRONTEND IMPACT: none required. The `id` field returned by /pending and
/approved is now the LINK's _id (not the vendor identity's _id) — this is
deliberate, so that Vendors.jsx's existing approve/reject/delete/deactivate
calls (which just pass back whatever `_id` they were given) transparently
operate on the correct per-tenant relationship without any frontend change.
RegisterVendor.jsx also needs no change — "vendor already registered" logic
now means "already has a link with THIS tenant", not "email exists anywhere".
"""

from fastapi import APIRouter, HTTPException, Request, BackgroundTasks
from fastapi.responses import StreamingResponse
from bson import ObjectId
from jose import jwt
from datetime import datetime, timedelta
from io import BytesIO
from html import escape
import re
from ..db import (
    vendors_collection,
    vendor_tenant_links_collection,
    product_mapping_collection,
    vendor_invites_collection,
    questionnaire_collection,
    tenants_collection,
)
from ..config import settings, frontend_url
from ..utils import hash_password, verify_password
from ..email_utils import (
    send_vendor_confirmation_email,
    send_vendor_invite_email,
    send_questionnaire_received_email,
)
from fastapi import Form, File, UploadFile, Depends, Header, Query
from typing import Dict, List, Optional
import cloudinary
import cloudinary.uploader

from .deps import get_hq_tenant
from ..db import admins_collection
from ..activity_log import log_activity
from ..vision_extract import extract_visiting_card

vendor_bp = APIRouter(prefix="/api/vendors", tags=["Vendors"])

SECRET_KEY = settings.secret_key
ALGORITHM = "HS256"

cloudinary.config(
    cloud_name=settings.cloudinary_cloud_name,
    api_key=settings.cloudinary_api_key,
    api_secret=settings.cloudinary_api_secret,
    secure=True,
)


def serialize_doc(doc):
    doc = dict(doc)
    doc["_id"] = str(doc["_id"])
    return doc


def _parse_brand_names(raw) -> List[str]:
    """A vendor can carry more than one brand (a distributor repping several
    lines, for instance). Accepts a list OR a comma-separated string so the
    SAME plain text input already used on the buyer's "Add Vendor" modal and
    the vendor's own registration form can express multiple brands without
    needing a dedicated tag-input widget on either side — the caller just
    types "Nike, Adidas, Puma"."""
    if isinstance(raw, list):
        return [b.strip() for b in raw if isinstance(b, str) and b.strip()]
    if isinstance(raw, str):
        return [b.strip() for b in raw.split(",") if b.strip()]
    return []


async def _require_active_vendor_invite(token: str) -> dict:
    """Return a usable invite; never trust client-side invite validation."""
    invite = await vendor_invites_collection.find_one({"token": token})
    if not invite:
        raise HTTPException(status_code=404, detail="Invite link is invalid.")

    expires_at = invite.get("expires_at")
    if invite.get("status") != "Pending":
        message = "This invite link has already been used." if invite.get("status") == "Registered" else "This invite link is no longer active."
        raise HTTPException(status_code=400, detail=message)
    if not expires_at or datetime.utcnow() > expires_at:
        await vendor_invites_collection.update_one(
            {"_id": invite["_id"], "status": "Pending"},
            {"$set": {"status": "Expired", "expired_at": datetime.utcnow()}},
        )
        raise HTTPException(status_code=400, detail="This invite link has expired.")
    return invite

def create_token(data: dict, expires_in: int = 3600):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(seconds=expires_in)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def decode_token(token: str):
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except jwt.ExpiredSignatureError:
        return None
    except jwt.JWTError:
        return None


async def require_vendor_identity(authorization: Optional[str] = Header(None)) -> dict:
    """Authenticate a vendor portal request and return its vendor identity."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization token missing")
    decoded = decode_token(authorization.split(" ", 1)[1])
    vendor_id = (decoded or {}).get("vendor_id")
    if not vendor_id or not ObjectId.is_valid(str(vendor_id)):
        raise HTTPException(status_code=401, detail="Invalid or expired vendor session")
    vendor = await vendors_collection.find_one({"_id": ObjectId(str(vendor_id))})
    if not vendor:
        raise HTTPException(status_code=401, detail="Vendor account not found")
    return vendor


def vendor_po_query(po_id: str, vendor_id: ObjectId) -> dict:
    if not ObjectId.is_valid(str(po_id)):
        raise HTTPException(status_code=400, detail="Invalid purchase order ID")
    return {
        "_id": ObjectId(str(po_id)),
        "$or": [{"vendor_id": vendor_id}, {"vendor_id": str(vendor_id)}],
    }

async def find_best_mapping(product_type: str, tenant_id: str):
    mapping = await product_mapping_collection.find_one({
        "tenant_id": tenant_id,
        "$expr": {"$eq": [{"$toLower": "$product_type"}, product_type.lower()]}
    })
    if mapping:
        return mapping
    mapping = await product_mapping_collection.find_one({
        "tenant_id": tenant_id,
        "product_type": {"$regex": product_type, "$options": "i"}
    })
    if mapping:
        return mapping
    words = product_type.lower().split()
    for w in words:
        mapping = await product_mapping_collection.find_one({
            "tenant_id": tenant_id,
            "product_type": {"$regex": w, "$options": "i"}
        })
        if mapping:
            return mapping
    return None


# ── Link/identity join helper ──────────────────────────────────────────────
# Every route that lists vendors for an HQ admin needs to show the SAME
# response shape the frontend already expects (name, brandName, email,
# status, source, vendor_code, etc.) — but that data now spans two
# documents. This merges a link doc with its identity doc into one flat
# dict, with "_id" set to the LINK's id (see FRONTEND IMPACT note above).

async def _merge_link_with_identity(link: dict, vendor: dict) -> dict:
    merged = {
        **{k: v for k, v in vendor.items() if k not in ("_id", "password")},
        **{k: v for k, v in link.items() if k != "_id"},
        "_id":        link["_id"],          # LINK id — what the frontend acts on
        "vendor_id":  str(vendor["_id"]),   # identity id, exposed for reference
    }
    return serialize_doc(merged)


# ---------------- Vendor APIs ----------------

@vendor_bp.post("/register")
async def register_vendor(request: Request, background_tasks: BackgroundTasks):
    """
    Register a new vendor — creates or reuses a vendor IDENTITY, and always
    creates a new relationship LINK for the resolved tenant.

    Invite-link registrations auto-approve immediately: the buyer already
    vetted this specific vendor by sending the invite, so a second manual
    "Approve" click after they fill the form is redundant friction — the
    vendor already set their password in this same request, so they can go
    straight to login. Self-registration (no invite, vendor picked a
    retailer off the public list) still needs manual HQ review, since there
    was no prior vetting.

    Public route — no HQ auth, since the vendor doesn't have an account yet.
    tenant_id CANNOT come from a JWT here. Two ways it can be resolved:

      1. INVITE-LINK path: token in the body → tenant_id comes from
         vendor_invites_collection.
      2. SELF-REGISTRATION path: no token → the frontend sends an explicit
         `tenant_id`, chosen by the vendor from the public retailer list
         (GET /api/tenants/public), validated against tenants_collection.

    MULTI-TENANT BEHAVIOR (new): if a vendor identity with this email
    already exists (e.g. they're already approved with Citimart and are
    now registering with Zudio), we do NOT reject with "already
    registered" — we reuse the existing identity and create a new link for
    the new tenant. Registration is only rejected if a link for THIS
    SPECIFIC (vendor, tenant) pair already exists.
    """
    body  = await request.json()
    email = (body.get("email") or "").strip().lower()
    token = (body.get("token") or "").strip()
    selected_tenant_id = (body.get("tenant_id") or "").strip()

    if not email:
        raise HTTPException(status_code=400, detail="Email is required.")

    # A plan chosen on the public pricing page is an onboarding request,
    # not an entitlement. Paid access activates only after approval and
    # the existing signed Razorpay subscription workflow.
    requested_plan = str(body.get("requested_plan") or "free").strip().lower()
    if requested_plan not in {"free", "standard", "premium"}:
        raise HTTPException(status_code=400, detail="Invalid requested plan.")

    # ── Resolve tenant_id — invite first, then explicit self-registration pick ──
    tenant_id = None
    source    = "self_registration"

    if token:
        invite = await _require_active_vendor_invite(token)
        invited_email = str(invite.get("email") or "").strip().lower()
        if invited_email and invited_email != email:
            raise HTTPException(status_code=400, detail="Use the email address that received this vendor invitation.")
        tenant_id = invite.get("tenant_id")
        source = "invite_link"
    if not tenant_id and selected_tenant_id:
        tenant = await tenants_collection.find_one({"tenant_id": selected_tenant_id})
        if not tenant:
            raise HTTPException(status_code=400, detail="Selected retailer was not found.")
        tenant_id = selected_tenant_id
        source    = "self_registration"

    if not tenant_id:
        raise HTTPException(
            status_code=400,
            detail="Please select a retailer to register with, or use a valid invite link. "
                   "Cannot determine which retailer this vendor is registering for."
        )

    # ── Resolve identity: reuse if this email already has one, else create ──
    identity = await vendors_collection.find_one({"email": email})

    # ⚠️ FIXED BUG: RegisterVendor.jsx's invite-link path collects a
    # password at registration time (its "Set Your Password" section only
    # renders when a token is present). This route previously discarded it
    # entirely — new identities always got password=None, password_set=
    # False, no matter what was submitted. That meant approve_vendor's
    # check ("does this identity already have a password?") always came
    # back False for invite-registered vendors too, so approval kept
    # sending a redundant "please set your password" email/prompt even
    # though the vendor had already set one. Now: if a password was
    # submitted, hash and store it immediately, and mark password_set=True
    # so approval correctly skips the redundant step.
    submitted_password = body.get("password")

    # Capture one primary classification at onboarding. Additional tags are
    # managed later from My Categories according to the subscription plan.
    raw_business_type = body.get("business_type") or []
    if isinstance(raw_business_type, str):
        raw_business_type = [raw_business_type]
    if not isinstance(raw_business_type, list):
        raise HTTPException(status_code=400, detail="business_type must be an array.")
    registration_business_types = list(dict.fromkeys(
        value.strip().lower() for value in raw_business_type
        if isinstance(value, str) and value.strip()
    ))
    invalid_business_types = set(registration_business_types) - VALID_BUSINESS_TYPES
    if invalid_business_types:
        raise HTTPException(status_code=400, detail=f"Invalid business type: {sorted(invalid_business_types)}")
    if len(registration_business_types) > 1:
        raise HTTPException(status_code=400, detail="Choose one primary business type during registration.")

    # Invite registration remains locked to its invitation tenant. Public
    # self-registration may request several retailer relationships at once.
    if token:
        tenant_ids = [tenant_id]
    else:
        raw_tenant_ids = body.get("tenant_ids") or [tenant_id]
        if not isinstance(raw_tenant_ids, list):
            raise HTTPException(status_code=400, detail="tenant_ids must be an array.")
        tenant_ids = list(dict.fromkeys(
            value.strip() for value in raw_tenant_ids
            if isinstance(value, str) and value.strip()
        ))
        if not tenant_ids:
            raise HTTPException(status_code=400, detail="Select at least one retailer.")
        valid_tenant_count = await tenants_collection.count_documents({
            "tenant_id": {"$in": tenant_ids},
            "status": {"$ne": "suspended"},
            "account_type": {"$ne": "single_store"},
        })
        if valid_tenant_count != len(tenant_ids):
            raise HTTPException(status_code=400, detail="One or more selected retailers are unavailable or are not eligible for vendor self-registration.")

    if identity:
        existing_link = await vendor_tenant_links_collection.find_one({
            "vendor_id": identity["_id"],
            "tenant_id": {"$in": tenant_ids},
        })
        if existing_link:
            raise HTTPException(
                status_code=400,
                detail="This vendor is already registered with one of the selected retailers."
            )
        vendor_id = identity["_id"]

        # Only fill in a password if this identity doesn't already have one.
        # Never overwrite an existing password from this unauthenticated
        # endpoint — that would let anyone re-register the same email under
        # a new invite/tenant and silently take over an existing login.
        if submitted_password and not identity.get("password_set"):
            await vendors_collection.update_one(
                {"_id": identity["_id"]},
                {"$set": {
                    "password":     hash_password(submitted_password),
                    "password_set": True,
                }}
            )
        if registration_business_types and not identity.get("business_type"):
            await vendors_collection.update_one(
                {"_id": identity["_id"]},
                {"$set": {"business_type": registration_business_types}},
            )
    else:
        identity_fields = [
            "name", "brandName", "companyType", "industryType",
            "ownerName", "contactName", "contactMobile", "email", "website",
            "address", "cityName", "state", "pincode", "pan",
            "gstCategory", "gstin", "gstState",
        ]
        identity_doc = {k: body.get(k) for k in identity_fields}
        identity_doc["email"] = email
        identity_doc["business_type"] = registration_business_types
        identity_doc["onboarding_requested_plan"] = requested_plan

        # brandName is a plain text field on the form, but can carry more
        # than one brand as a comma-separated list ("Nike, Adidas, Puma") —
        # normalize it into brandNames here so anything that wants to work
        # with them individually can, while brandName itself stays a clean,
        # consistently-joined string for every existing display fallback
        # elsewhere in the app.
        identity_doc["brandNames"] = _parse_brand_names(identity_doc.get("brandName"))
        identity_doc["brandName"] = ", ".join(identity_doc["brandNames"])

        if submitted_password:
            identity_doc["password"]     = hash_password(submitted_password)
            identity_doc["password_set"] = True
        else:
            identity_doc["password"]     = None
            identity_doc["password_set"] = False

        identity_doc["created_at"] = datetime.utcnow()
        result = await vendors_collection.insert_one(identity_doc)
        vendor_id = result.inserted_id

    # ── Create the per-tenant relationship link ───────────────────────────
    link_docs = [{
        "vendor_id": vendor_id,
        "tenant_id": selected_id,
        "product_type": body.get("product_type", ""),
        "division": None, "section": None, "department": None,
        "status": "Pending", "source": source,
        "requested_plan": requested_plan,
        "created_at": datetime.utcnow(),
    } for selected_id in tenant_ids]
    link_result = await vendor_tenant_links_collection.insert_many(link_docs)
    link_ids = [str(link_id) for link_id in link_result.inserted_ids]

    # Invite-link registrations auto-approve — see the docstring above.
    # tenant_ids is always exactly one entry on this path (locked to the
    # invitation's own tenant), so there's exactly one link to finalize.
    approval = None
    if source == "invite_link":
        vendor = await vendors_collection.find_one({"_id": vendor_id})
        link = await vendor_tenant_links_collection.find_one({"_id": link_result.inserted_ids[0]})
        if vendor and link:
            approval = await _finalize_vendor_approval(
                link, vendor, tenant_id, body.get("product_type", ""),
                approver_name="System", approver_role="Auto-Approval", approver_department="Invite",
                background_tasks=background_tasks,
            )

    return {
        "message":    "Vendor registered and approved — you can log in now." if approval else "Vendor registered successfully",
        "status":     "Approved" if approval else "Pending",
        "vendor_id":  str(vendor_id),
        "link_id":    link_ids[0],
        "link_ids":   link_ids,
        "tenant_ids": tenant_ids,
        "business_type": registration_business_types,
        "requested_plan": requested_plan,
        "vendor_code": (approval or {}).get("vendor_code"),
    }


@vendor_bp.get("/pending")
async def get_pending_vendors(ctx: dict = Depends(get_hq_tenant)):
    """List all vendor relationships Pending for this tenant."""
    rows = []
    async for link in vendor_tenant_links_collection.find({"status": "Pending", "tenant_id": ctx["tenant_id"]}):
        vendor = await vendors_collection.find_one({"_id": link["vendor_id"]})
        if vendor:
            rows.append(await _merge_link_with_identity(link, vendor))
    return rows


@vendor_bp.get("/approved")
async def get_approved_vendors(ctx: dict = Depends(get_hq_tenant)):
    """List all vendor relationships Approved for this tenant."""
    rows = []
    async for link in vendor_tenant_links_collection.find({"status": "Approved", "tenant_id": ctx["tenant_id"]}):
        vendor = await vendors_collection.find_one({"_id": link["vendor_id"]})
        if vendor:
            rows.append(await _merge_link_with_identity(link, vendor))
    return rows


async def _finalize_vendor_approval(
    link: dict, vendor: dict, tenant_id: str, product_type: str,
    approver_name: str, approver_role: str, approver_department: str,
    background_tasks: BackgroundTasks,
) -> dict:
    """Shared by the manual HQ approve_vendor() below and register_vendor()'s
    invite-link auto-approval — vendor code assignment, division/section/
    department mapping, and the "you're in" email are identical either way,
    only who/what triggered the approval differs."""
    mapping = await product_mapping_collection.find_one(
        {"tenant_id": tenant_id, "product_type": {"$regex": product_type or link.get("product_type", ""), "$options": "i"}}
    )
    if mapping:
        division, section, department = mapping.get("division"), mapping.get("section"), mapping.get("department")
    else:
        division, section, department = "Uncategorized", "-", "-"

    # Vendor code — scoped per tenant, so numbering is independent per retailer
    last_link = await vendor_tenant_links_collection.find_one(
        {"vendor_code": {"$exists": True}, "tenant_id": tenant_id},
        sort=[("vendor_code", -1)]
    )
    if last_link and "vendor_code" in last_link:
        try:
            new_num = int(last_link["vendor_code"].split("-")[1]) + 1
        except Exception:
            new_num = 1
    else:
        new_num = 1
    vendor_code = f"VEN-{new_num:05d}"

    await vendor_tenant_links_collection.update_one(
        {"_id": link["_id"]},
        {
            "$set": {
                "status":      "Approved",
                "division":    division,
                "section":     section,
                "department":  department,
                "vendor_code": vendor_code,
                "approved_at": datetime.utcnow(),
                "kyb_status": "Not started",
                "kyb_required_after": datetime.utcnow(),
            },
            "$push": {
                "approvals": {
                    "role": approver_role, "approved_by": approver_name,
                    "department": approver_department, "time": datetime.utcnow(),
                }
            },
        }
    )

    # Password-setup email only needed if this identity has no password yet.
    # An invite-link registrant already set one during registration itself,
    # so this branch only ever fires for the manual-approval path (a
    # self-registered or pre-password-era vendor).
    if not vendor.get("password_set"):
        setup_token = create_token(
            {"vendor_id": str(vendor["_id"]), "email": vendor["email"]},
            expires_in=604800,
        )
        setup_link = frontend_url(f'merchandiser-seller/setup-password?token={setup_token}')
        background_tasks.add_task(
            send_vendor_confirmation_email,
            vendor["email"], vendor.get("name", ""), vendor.get("brandName", "Your Brand"), setup_link,
        )
        email_note = "Confirmation email sent (valid 7 days) — please set your password to log in."
    else:
        email_note = f"{vendor.get('name','This vendor')} already has login access; this retailer relationship is now live on their next login."

    await log_activity(
        approver_name, f"Approved vendor {vendor.get('name', vendor_code)} ({vendor_code})", type="create",
    )

    return {
        "vendor_code": vendor_code, "division": division, "section": section,
        "department": department, "email_note": email_note,
    }


@vendor_bp.post("/approve/{link_id}")
async def approve_vendor(
    link_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    ctx: dict = Depends(get_hq_tenant),
):
    """
    Approve a vendor-tenant relationship. `link_id` is the LINK's _id (see
    module docstring — /pending and /approved return link ids as "_id" so
    the existing frontend needs no change).

    Vendor code, division/section/department mapping, and the confirmation
    email are all scoped to THIS relationship — approving a vendor at
    Citimart has no effect on their (possibly still-Pending, or
    nonexistent) relationship with any other retailer.
    """
    body = await request.json()
    product_type = body.get("product_type")

    try:
        link = await vendor_tenant_links_collection.find_one({
            "_id": ObjectId(link_id),
            "tenant_id": ctx["tenant_id"],
        })
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid vendor ID")

    if not link:
        raise HTTPException(status_code=404, detail="Vendor not found")

    if link.get("status") == "Approved":
        raise HTTPException(status_code=400, detail="Vendor is already approved.")

    vendor = await vendors_collection.find_one({"_id": link["vendor_id"]})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor identity not found")

    approver_name = "Unknown User"
    approver_role = "Unknown Role"
    approver_department = ctx.get("department", "Unknown Department")
    if ctx.get("admin_id"):
        approver = await admins_collection.find_one({"_id": ObjectId(ctx["admin_id"])})
        if approver:
            approver_name = approver.get("name", approver_name)
            approver_department = approver.get("department", approver_department)
            approver_role = approver.get("department", approver_role)

    result = await _finalize_vendor_approval(
        link, vendor, ctx["tenant_id"], product_type,
        approver_name, approver_role, approver_department, background_tasks,
    )

    return {
        "message": f"✅ Vendor {result['vendor_code']} approved by {approver_name} ({approver_role}, {approver_department}). {result['email_note']}",
        "vendor_code": result["vendor_code"],
        "division": result["division"], "section": result["section"], "department": result["department"],
        "approved_by": approver_name, "approved_role": approver_role, "approved_department": approver_department,
    }


@vendor_bp.post("/reject/{link_id}")
async def reject_vendor(link_id: str, ctx: dict = Depends(get_hq_tenant)):
    """Reject a pending vendor-tenant relationship (link-scoped, not identity-wide)."""
    if not ObjectId.is_valid(link_id):
        raise HTTPException(status_code=400, detail="Invalid vendor ID")

    link = await vendor_tenant_links_collection.find_one({"_id": ObjectId(link_id), "tenant_id": ctx["tenant_id"]})
    if not link:
        raise HTTPException(status_code=404, detail="Vendor not found")
    if link.get("status") == "Approved":
        raise HTTPException(status_code=400, detail="Cannot reject an already-approved vendor.")

    await vendor_tenant_links_collection.update_one(
        {"_id": ObjectId(link_id), "tenant_id": ctx["tenant_id"]},
        {"$set": {"status": "Rejected", "rejected_at": datetime.utcnow(), "rejected_by": ctx.get("admin_id")}}
    )

    rejected_vendor = await vendors_collection.find_one({"_id": link["vendor_id"]}, {"name": 1, "vendor_name": 1})
    vendor_label = (rejected_vendor or {}).get("name") or (rejected_vendor or {}).get("vendor_name") or str(link["vendor_id"])
    await log_activity(ctx.get("admin_name", ""), f"Rejected vendor {vendor_label}", type="warning")

    return {"message": "Vendor rejected successfully."}


@vendor_bp.delete("/delete/{link_id}")
async def delete_vendor(link_id: str, ctx: dict = Depends(get_hq_tenant)):
    """
    Delete a vendor-tenant relationship. Only removes THIS tenant's
    relationship — the vendor's identity (and any other retailer's
    relationship with them) is untouched. If this was their only
    relationship anywhere, the identity document is left in place
    (harmless orphan; they simply won't appear in any tenant's list) —
    deliberately not cascading a delete into vendor identity, since that
    would risk deleting a login another tenant still depends on if this
    check raced with a concurrent registration. Safe default: leave it.
    """
    if not ObjectId.is_valid(link_id):
        raise HTTPException(status_code=400, detail="Invalid vendor ID")

    result = await vendor_tenant_links_collection.delete_one({"_id": ObjectId(link_id), "tenant_id": ctx["tenant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor deleted successfully."}


@vendor_bp.post("/deactivate/{link_id}")
async def deactivate_vendor(link_id: str, ctx: dict = Depends(get_hq_tenant)):
    """Deactivate an approved vendor-tenant relationship (link-scoped)."""
    if not ObjectId.is_valid(link_id):
        raise HTTPException(status_code=400, detail="Invalid vendor ID")

    link = await vendor_tenant_links_collection.find_one({"_id": ObjectId(link_id), "tenant_id": ctx["tenant_id"]})
    if not link:
        raise HTTPException(status_code=404, detail="Vendor not found")

    await vendor_tenant_links_collection.update_one(
        {"_id": ObjectId(link_id), "tenant_id": ctx["tenant_id"]},
        {"$set": {"status": "Deactivated", "deactivated_at": datetime.utcnow(), "deactivated_by": ctx.get("admin_id")}}
    )
    return {"message": "Vendor deactivated successfully."}


@vendor_bp.post("/setup-password")
async def setup_vendor_password(request: Request):
    """Set vendor password via confirmation link. Public — sets password on the IDENTITY, not any one link."""
    body = await request.json()
    token = body.get("token")
    password = body.get("password")

    decoded = decode_token(token)
    if not decoded:
        raise HTTPException(status_code=400, detail="Invalid or expired token")

    vendor_id = decoded.get("vendor_id")
    hashed = hash_password(password)

    await vendors_collection.update_one(
        {"_id": ObjectId(vendor_id)},
        {"$set": {"password": hashed, "password_set": True}},
    )
    return {"message": "Password setup successful, please login now."}


@vendor_bp.post("/login")
async def vendor_login(request: Request):
    """
    Vendor login — identity-level, not tenant-scoped. A vendor with
    relationships to multiple retailers logs in ONCE; which retailers they
    can act on is discovered via GET /api/vendors/my-tenant after login,
    not baked into the JWT. The JWT therefore carries vendor_id + email
    only — same shape as before this change, so /me, /my-purchaseorders
    etc. that decode vendor_id from the token are unaffected.

    Login succeeds if the identity has a password set AND at least one
    Approved relationship exists anywhere (being Pending/Rejected
    everywhere means there's nothing to log in and do yet).
    """
    body = await request.json()
    email = (body.get("email") or "").strip().lower()
    password = body.get("password")

    vendor = await vendors_collection.find_one({"email": email})
    if not vendor:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    approved_links = await vendor_tenant_links_collection.find({
        "vendor_id": vendor["_id"], "status": "Approved",
    }).to_list(length=None)
    if not approved_links:
        raise HTTPException(status_code=403, detail="Vendor not approved yet by any retailer.")

    if not verify_password(password, vendor.get("password", "")):
        raise HTTPException(status_code=401, detail="Incorrect password")

    token = create_token({"vendor_id": str(vendor["_id"]), "email": vendor["email"]})
    for link in approved_links:
        tenant = await tenants_collection.find_one({"tenant_id": link.get("tenant_id")}, {"company_name": 1})
        await log_activity(
            vendor.get("name") or vendor.get("vendor_name") or vendor.get("email", ""),
            "Vendor logged in", type="info",
            tenant_id=link.get("tenant_id"),
            tenant_name=(tenant or {}).get("company_name") or link.get("tenant_id"),
            actor_email=vendor.get("email"), actor_role="Vendor",
        )
    return {
        "access_token": token,
        "vendor_id":    str(vendor["_id"]),
        "vendor_name":  vendor.get("name") or vendor.get("vendor_name") or "",
        "email":        vendor.get("email", ""),
        "redirect":     "/merchandiser-seller",
    }


@vendor_bp.get("/me")
async def get_vendor_profile(authorization: str = Header(None)):
    """Fetch logged-in vendor's IDENTITY profile. Unchanged in shape — no tenant/status fields here anymore (see /my-tenant for those, per relationship)."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization token missing")

    token = authorization.split(" ")[1]
    decoded = decode_token(token)
    if not decoded:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    vendor_id = decoded.get("vendor_id")
    if not vendor_id:
        raise HTTPException(status_code=400, detail="Invalid token payload")

    vendor = await vendors_collection.find_one({"_id": ObjectId(vendor_id)})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    vendor["_id"] = str(vendor["_id"])
    vendor.pop("password", None)
    return vendor


@vendor_bp.patch("/me/settings")
async def update_vendor_settings(request: Request, vendor: dict = Depends(require_vendor_identity)):
    """Update safe vendor profile fields and vendor-portal preferences."""
    body = await request.json()
    profile = body.get("profile") or {}
    preferences = body.get("preferences") or {}
    if not isinstance(profile, dict) or not isinstance(preferences, dict):
        raise HTTPException(status_code=400, detail="profile and preferences must be objects.")

    updates = {}
    profile_limits = {
        "name": 160,
        "contactMobile": 40,
        "address": 500,
        "city": 120,
        "website": 255,
        # Tax/registration details — deliberately NOT collected on the public
        # registration form (a brand-new vendor won't hand PAN/GST to a site
        # they've never heard of before they've even logged in). Completed
        # here instead, once they're inside their own dashboard.
        "pan": 10,
        "gstin": 15,
        "gstCategory": 40,
        "gstState": 60,
    }
    for field, maximum in profile_limits.items():
        if field not in profile:
            continue
        value = str(profile.get(field) or "").strip().upper() if field in ("pan", "gstin") else str(profile.get(field) or "").strip()
        if len(value) > maximum:
            raise HTTPException(status_code=400, detail=f"{field} must be {maximum} characters or fewer.")
        if field == "name" and not value:
            raise HTTPException(status_code=400, detail="Business name cannot be empty.")
        if field == "pan" and value and not re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", value):
            raise HTTPException(status_code=400, detail="PAN must be in the format AAAAA9999A.")
        if field == "gstin" and value and not re.fullmatch(r"\d{2}[A-Z]{5}\d{4}[A-Z][A-Z\d]Z[A-Z\d]", value):
            raise HTTPException(status_code=400, detail="GSTIN must be a valid 15-character GST number.")
        updates[field] = value

    current_settings = vendor.get("settings") if isinstance(vendor.get("settings"), dict) else {}
    default_notifications = {
        "purchase_orders": True,
        "rfqs_and_messages": True,
        "supplier_returns": True,
        "email_alerts": True,
        "whatsapp_alerts": False,
    }
    default_orders = {
        "default_lead_time_days": 7,
        "minimum_order_quantity": 1,
        "default_payment_terms": "",
        "return_policy": "",
    }
    notifications = {**default_notifications, **(current_settings.get("notification_preferences") or {})}
    order_defaults = {**default_orders, **(current_settings.get("order_preferences") or {})}

    requested_notifications = preferences.get("notifications") or {}
    if not isinstance(requested_notifications, dict):
        raise HTTPException(status_code=400, detail="Notification preferences must be an object.")
    for field in default_notifications:
        if field in requested_notifications:
            notifications[field] = bool(requested_notifications[field])

    requested_orders = preferences.get("order_preferences") or {}
    if not isinstance(requested_orders, dict):
        raise HTTPException(status_code=400, detail="Order preferences must be an object.")
    for field in ("default_lead_time_days", "minimum_order_quantity"):
        if field in requested_orders:
            try:
                value = int(requested_orders[field])
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"{field} must be a whole number.")
            if value < 0 or value > 3650:
                raise HTTPException(status_code=400, detail=f"{field} must be between 0 and 3650.")
            order_defaults[field] = value
    for field, maximum in (("default_payment_terms", 160), ("return_policy", 1000)):
        if field in requested_orders:
            value = str(requested_orders[field] or "").strip()
            if len(value) > maximum:
                raise HTTPException(status_code=400, detail=f"{field} must be {maximum} characters or fewer.")
            order_defaults[field] = value

    updates["settings.notification_preferences"] = notifications
    updates["settings.order_preferences"] = order_defaults
    updates["settings.updated_at"] = datetime.utcnow()
    await vendors_collection.update_one({"_id": vendor["_id"]}, {"$set": updates})
    updated = await vendors_collection.find_one({"_id": vendor["_id"]})
    return {
        "message": "Vendor settings saved.",
        "data": {
            "name": updated.get("name", ""),
            "email": updated.get("email", ""),
            "contactMobile": updated.get("contactMobile", ""),
            "address": updated.get("address", ""),
            "city": updated.get("city", ""),
            "website": updated.get("website", ""),
            "pan": updated.get("pan", ""),
            "gstin": updated.get("gstin", ""),
            "gstCategory": updated.get("gstCategory", ""),
            "gstState": updated.get("gstState", ""),
            "settings": {
                "notification_preferences": notifications,
                "order_preferences": order_defaults,
            },
        },
    }

@vendor_bp.post("/me/kyb/documents/{document_type}")
async def upload_vendor_kyb_document(
    document_type: str,
    file: UploadFile = File(...),
    vendor: dict = Depends(require_vendor_identity),
):
    """Upload a vendor KYB image/PDF and return its HTTPS storage URL."""
    field_by_type = {
        "gst_certificate": "gst_certificate_url",
        "pan_document": "pan_document_url",
        "cancelled_cheque": "cancelled_cheque_url",
    }
    if document_type not in field_by_type:
        raise HTTPException(status_code=400, detail="Choose a valid KYB document type.")
    allowed_types = {"image/jpeg", "image/png", "image/webp", "application/pdf"}
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Upload JPG, PNG, WEBP or PDF only.")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="The selected file is empty.")
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="KYB documents must be 10 MB or smaller.")
    try:
        result = cloudinary.uploader.upload(
            raw,
            folder=f"rms/vendor-kyb/{vendor['_id']}",
            resource_type="auto",
            public_id=f"{document_type}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
            use_filename=True,
            unique_filename=True,
        )
    except Exception as exc:
        raise HTTPException(status_code=502, detail="Could not upload the document. Please try again.") from exc
    url = result.get("secure_url")
    if not url:
        raise HTTPException(status_code=502, detail="Document storage did not return a secure URL.")
    await vendors_collection.update_one(
        {"_id": vendor["_id"]},
        {"$set": {f"kyb.{field_by_type[document_type]}": url, "kyb.updated_at": datetime.utcnow()}},
    )
    return {"url": url, "name": file.filename, "content_type": file.content_type}

@vendor_bp.get("/me/kyb")
async def get_vendor_kyb(vendor: dict = Depends(require_vendor_identity)):
    """Vendor-facing KYB record. Bank account numbers are never returned."""
    kyb = dict(vendor.get("kyb") or {})
    kyb.pop("account_number", None)
    links = await vendor_tenant_links_collection.find(
        {"vendor_id": vendor["_id"]}, {"tenant_id": 1, "status": 1, "kyb_status": 1, "kyb_note": 1, "kyb_reviewed_at": 1}
    ).to_list(length=None)
    return {"data": {
        "legal_name": kyb.get("legal_name", ""), "business_address": kyb.get("business_address", ""),
        "bank_account_holder": kyb.get("bank_account_holder", ""), "bank_name": kyb.get("bank_name", ""),
        "ifsc": kyb.get("ifsc", ""), "account_last4": kyb.get("account_last4", ""),
        "gst_certificate_url": kyb.get("gst_certificate_url", ""), "pan_document_url": kyb.get("pan_document_url", ""),
        "cancelled_cheque_url": kyb.get("cancelled_cheque_url", ""), "submitted_at": kyb.get("submitted_at"),
        "relationships": [{"tenant_id": row.get("tenant_id"), "relationship_status": row.get("status", "Pending"), "status": row.get("kyb_status", "Not started"), "note": row.get("kyb_note", ""), "reviewed_at": row.get("kyb_reviewed_at")} for row in links],
    }}


@vendor_bp.patch("/me/kyb")
async def submit_vendor_kyb(request: Request, vendor: dict = Depends(require_vendor_identity)):
    """Store the vendor's KYB submission; raw account numbers are discarded after masking."""
    body = await request.json()
    required = ("legal_name", "business_address", "bank_account_holder", "bank_name", "ifsc", "account_number")
    values = {key: str(body.get(key) or "").strip() for key in required}
    missing = [key.replace("_", " ") for key, value in values.items() if not value]
    if missing:
        raise HTTPException(status_code=400, detail=f"Complete: {', '.join(missing)}.")
    if not vendor.get("pan") or not vendor.get("gstin"):
        raise HTTPException(status_code=400, detail="Save valid PAN and GSTIN in Tax & registration before submitting KYB.")
    account_number = re.sub(r"[\s-]", "", values["account_number"])
    if not re.fullmatch(r"\d{9,18}", account_number):
        raise HTTPException(status_code=400, detail="Bank account number must contain 9 to 18 digits.")
    ifsc = values["ifsc"].upper()
    if not re.fullmatch(r"[A-Z]{4}0[A-Z0-9]{6}", ifsc):
        raise HTTPException(status_code=400, detail="IFSC must be in the format AAAA0AAAAAA.")
    urls = {}
    for key in ("gst_certificate_url", "pan_document_url", "cancelled_cheque_url"):
        value = str(body.get(key) or "").strip()
        if value and not re.match(r"^https://", value, re.I):
            raise HTTPException(status_code=400, detail=f"{key.replace('_', ' ')} must be a secure https link.")
        urls[key] = value
    kyb = {
        "legal_name": values["legal_name"][:200], "business_address": values["business_address"][:600],
        "bank_account_holder": values["bank_account_holder"][:160], "bank_name": values["bank_name"][:160],
        "ifsc": ifsc, "account_last4": account_number[-4:], **urls,
        "submitted_at": datetime.utcnow(), "updated_at": datetime.utcnow(),
    }
    await vendors_collection.update_one({"_id": vendor["_id"]}, {"$set": {"kyb": kyb}})
    await vendor_tenant_links_collection.update_many(
        {"vendor_id": vendor["_id"], "status": "Approved", "kyb_status": {"$ne": "Verified"}},
        {"$set": {"kyb_status": "Submitted", "kyb_submitted_at": datetime.utcnow(), "kyb_note": "Awaiting retailer finance verification."}},
    )
    return {"message": "KYB submitted. Each connected retailer can now review it.", "account_last4": account_number[-4:]}


@vendor_bp.get("/kyb/{link_id}")
async def get_vendor_kyb_for_review(link_id: str, ctx: dict = Depends(get_hq_tenant)):
    if not ObjectId.is_valid(link_id):
        raise HTTPException(status_code=400, detail="Invalid vendor relationship.")
    link = await vendor_tenant_links_collection.find_one({"_id": ObjectId(link_id), "tenant_id": ctx["tenant_id"]})
    if not link:
        raise HTTPException(status_code=404, detail="Vendor relationship not found.")
    vendor = await vendors_collection.find_one({"_id": link["vendor_id"]}, {"password": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor identity not found.")
    kyb = dict(vendor.get("kyb") or {})
    kyb.pop("account_number", None)
    return {"data": {"vendor": {"name": vendor.get("name") or vendor.get("vendor_name"), "email": vendor.get("email"), "pan": vendor.get("pan"), "gstin": vendor.get("gstin")}, "kyb": kyb, "status": link.get("kyb_status", "Not started"), "note": link.get("kyb_note", "")}}


@vendor_bp.patch("/kyb/{link_id}/review")
async def review_vendor_kyb(link_id: str, request: Request, ctx: dict = Depends(get_hq_tenant)):
    if not ObjectId.is_valid(link_id):
        raise HTTPException(status_code=400, detail="Invalid vendor relationship.")
    body = await request.json()
    status = str(body.get("status") or "").strip()
    note = str(body.get("note") or "").strip()[:1000]
    if status not in {"Verified", "Needs changes"}:
        raise HTTPException(status_code=400, detail="Choose Verified or Needs changes.")
    link = await vendor_tenant_links_collection.find_one({"_id": ObjectId(link_id), "tenant_id": ctx["tenant_id"]})
    if not link:
        raise HTTPException(status_code=404, detail="Vendor relationship not found.")
    await vendor_tenant_links_collection.update_one({"_id": link["_id"]}, {"$set": {"kyb_status": status, "kyb_note": note, "kyb_reviewed_at": datetime.utcnow(), "kyb_reviewed_by": ctx.get("admin_id")}})
    await log_activity(ctx.get("admin_name", "HQ Admin"), f"{status} vendor KYB", type="info", tenant_id=ctx["tenant_id"], actor_role="HQ Admin")
    return {"message": f"Vendor KYB marked {status}."}

VALID_BUSINESS_TYPES = {
    "general_vendor", "wholesaler", "manufacturer", "retailer",
    "fabric_supplier", "exporter", "distributor", "job_worker",
}


@vendor_bp.patch("/me/classification")
async def update_vendor_classification(request: Request, authorization: str = Header(None)):
    """
    Vendor self-service: sets which business_type(s) they are (general vendor /
    wholesaler / manufacturer / retailer / fabric supplier / exporter /
    distributor — multi-select)
    and free-text product_categories (e.g. "casual t-shirts", "formal
    shirts"). This is what lets a buyer SEARCH vendors by category instead
    of only browsing ones they already have a relationship with — see
    catalogue_routes.py's vendor/{vendor_id} route, which is the
    per-vendor catalogue view this classification feeds into for discovery.

    Tier-limited: the number of business_type tags a vendor can hold is
    capped by their subscription (Free=1, Standard=3, Premium=unlimited —
    see subscription_routes.TIER_CONFIG). Imported lazily inside the
    function body rather than at module load time, to avoid a circular
    import — subscription_routes.py itself imports decode_token from this
    file.
    """
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization token missing")
    decoded = decode_token(authorization.split(" ")[1])
    if not decoded:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    vendor_id = decoded.get("vendor_id")
    if not vendor_id:
        raise HTTPException(status_code=400, detail="Invalid token payload")

    body = await request.json()
    business_type = body.get("business_type", [])
    product_categories = body.get("product_categories", [])

    if not isinstance(business_type, list) or not isinstance(product_categories, list):
        raise HTTPException(status_code=400, detail="business_type and product_categories must both be arrays.")

    invalid = set(business_type) - VALID_BUSINESS_TYPES
    if invalid:
        raise HTTPException(status_code=400, detail=f"Invalid business_type value(s): {sorted(invalid)}. Valid: {sorted(VALID_BUSINESS_TYPES)}")

    from .subscription_routes import get_vendor_tier  # lazy import — see docstring
    tier = await get_vendor_tier(vendor_id)
    limit = tier["business_type_limit"]
    if limit is not None and len(business_type) > limit:
        raise HTTPException(
            status_code=403,
            detail=f"Your {tier['label']} plan allows up to {limit} business type tag(s). "
                   f"You selected {len(business_type)}. Upgrade to tag more."
        )

    clean_categories = [c.strip() for c in product_categories if isinstance(c, str) and c.strip()][:50]

    await vendors_collection.update_one(
        {"_id": ObjectId(vendor_id)},
        {"$set": {
            "business_type":      business_type,
            "product_categories": clean_categories,
            "classification_updated_at": datetime.utcnow(),
        }}
    )
    return {"status": "success", "message": "Classification updated.", "business_type": business_type, "product_categories": clean_categories}


@vendor_bp.get("/my-tenant")
async def get_my_tenant(authorization: str = Header(None)):
    """
    Vendor-facing: returns EVERY retailer this vendor has a relationship
    with — genuinely plural now. Powers the "Retailers" tab in the vendor
    dashboard: one row per (vendor, tenant) link, each with its own status,
    vendor_code, source, and approval date, independent of the others.
    """
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization token missing")

    token = authorization.split(" ")[1]
    decoded = decode_token(token)
    if not decoded:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    vendor_id = decoded.get("vendor_id")
    if not vendor_id:
        raise HTTPException(status_code=400, detail="Invalid token payload")

    try:
        vendor_oid = ObjectId(vendor_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid token payload")

    vendor = await vendors_collection.find_one({"_id": vendor_oid})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    retailers = []
    async for link in vendor_tenant_links_collection.find({"vendor_id": vendor_oid}).sort("created_at", -1):
        tenant = await tenants_collection.find_one({"tenant_id": link.get("tenant_id")})
        retailers.append({
            "tenant_id":    link.get("tenant_id"),
            "company_name": (tenant or {}).get("company_name") or link.get("tenant_id"),
            "source":       link.get("source", ""),
            "vendor_code":  link.get("vendor_code", ""),
            "status":       link.get("status", ""),
            "division":     link.get("division"),
            "section":      link.get("section"),
            "department":   link.get("department"),
            "approved_at":  str(link.get("approved_at")) if link.get("approved_at") else None,
            "created_at":   str(link.get("created_at")) if link.get("created_at") else None,
        })

    return {"status": "success", "data": retailers}


# ------------------- VENDOR PURCHASE ORDER ROUTES -------------------
from app.db import purchaseorders_collection, product_collection

def _vendor_dashboard_date(value):
    """Return a naive UTC datetime for dashboard month comparisons."""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, str):
        for date_format in ("%Y-%m-%d", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(value[:19], date_format)
            except ValueError:
                continue
    return None


@vendor_bp.get("/dashboard-summary")
async def get_vendor_dashboard_summary(authorization: str = Header(None)):
    """Return real, vendor-scoped headline metrics for the vendor dashboard."""
    vendor = await require_vendor_identity(authorization)
    vendor_oid = vendor["_id"]
    vendor_filter = {"$or": [{"vendor_id": vendor_oid}, {"vendor_id": str(vendor_oid)}]}

    now = datetime.utcnow()
    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if this_month_start.month == 1:
        previous_month_start = this_month_start.replace(year=this_month_start.year - 1, month=12)
    else:
        previous_month_start = this_month_start.replace(month=this_month_start.month - 1)

    total_products = await product_collection.count_documents(vendor_filter)
    products_added_this_month = await product_collection.count_documents({
        "$and": [
            vendor_filter,
            {"$or": [
                {"createdAt": {"$gte": this_month_start}},
                {"created_at": {"$gte": this_month_start}},
            ]},
        ],
    })

    orders = await purchaseorders_collection.find(
        vendor_filter,
        {"status": 1, "netAmount": 1, "createdAt": 1, "orderDate": 1, "currency": 1, "items": 1},
    ).to_list(None)

    pending_statuses = {"Pending", "SentToVendor", "WalkinAccepted"}
    confirmed_statuses = {"Approved", "PartiallyReceived", "Received", "GRNCompleted", "Paid"}
    pending_orders = sum(1 for order in orders if order.get("status") in pending_statuses)

    confirmed_revenue = 0.0
    current_month_revenue = 0.0
    previous_month_revenue = 0.0
    currencies = set()
    status_counts: dict[str, int] = {}
    # Last 6 calendar months including the current one, oldest first.
    month_buckets: list[tuple[int, int]] = []
    cursor_year, cursor_month = this_month_start.year, this_month_start.month
    for _ in range(6):
        month_buckets.append((cursor_year, cursor_month))
        cursor_month -= 1
        if cursor_month == 0:
            cursor_month = 12
            cursor_year -= 1
    month_buckets.reverse()
    trend_map = {f"{y:04d}-{m:02d}": 0.0 for y, m in month_buckets}
    item_totals: dict[str, dict] = {}

    for order in orders:
        status = order.get("status") or "Unknown"
        status_counts[status] = status_counts.get(status, 0) + 1

        if status not in confirmed_statuses:
            continue
        try:
            amount = float(order.get("netAmount") or 0)
        except (TypeError, ValueError):
            amount = 0.0
        confirmed_revenue += amount
        currencies.add(order.get("currency") or "INR")

        order_date = _vendor_dashboard_date(order.get("createdAt")) or _vendor_dashboard_date(order.get("orderDate"))
        if order_date:
            bucket_key = f"{order_date.year:04d}-{order_date.month:02d}"
            if bucket_key in trend_map:
                trend_map[bucket_key] += amount
            if order_date >= this_month_start:
                current_month_revenue += amount
            elif order_date >= previous_month_start:
                previous_month_revenue += amount

        for item in order.get("items") or []:
            description = str(item.get("description") or "Item").strip() or "Item"
            qty = float(item.get("quantity") or 0)
            rate = float(item.get("rate") or 0)
            bucket = item_totals.setdefault(description, {"description": description, "quantity": 0.0, "revenue": 0.0})
            bucket["quantity"] += qty
            bucket["revenue"] += qty * rate

    revenue_change_pct = None
    if previous_month_revenue > 0:
        revenue_change_pct = round(
            ((current_month_revenue - previous_month_revenue) / previous_month_revenue) * 100,
            1,
        )

    monthly_revenue_trend = [
        {"month": key, "revenue": round(value, 2)}
        for key, value in trend_map.items()
    ]
    order_status_breakdown = [
        {"status": status, "count": count}
        for status, count in sorted(status_counts.items(), key=lambda pair: pair[1], reverse=True)
    ]
    top_items = sorted(item_totals.values(), key=lambda row: row["revenue"], reverse=True)[:5]
    for row in top_items:
        row["quantity"] = round(row["quantity"], 2)
        row["revenue"] = round(row["revenue"], 2)

    return {
        "total_products": total_products,
        "products_added_this_month": products_added_this_month,
        "pending_orders": pending_orders,
        "confirmed_revenue": round(confirmed_revenue, 2),
        "current_month_revenue": round(current_month_revenue, 2),
        "previous_month_revenue": round(previous_month_revenue, 2),
        "revenue_change_pct": revenue_change_pct,
        "currency": currencies.pop() if len(currencies) == 1 else "INR",
        "monthly_revenue_trend": monthly_revenue_trend,
        "order_status_breakdown": order_status_breakdown,
        "top_items": top_items,
    }

@vendor_bp.get("/my-purchaseorders")
async def get_my_purchase_orders(authorization: str = Header(None)):
    """
    Vendor fetches their own POs across ALL retailer relationships. Each PO
    already carries its own tenant_id from purchaseorder_routes.py's
    tenant-scoping — this route doesn't need to filter by relationship
    status itself, since a PO could only ever have been created against a
    vendor_id in the first place by an HQ admin at that specific tenant.
    A vendor now correctly sees POs from every retailer they work with in
    one unified list, which is a natural benefit of the identity/link
    split rather than something that needed extra code here.
    """
    vendor = await require_vendor_identity(authorization)
    vendor_oid  = vendor["_id"]

    tenant_names = {}

    cursor = purchaseorders_collection.find({
        "$or": [
            {"vendor_id": vendor_oid},
            {"vendor_id": str(vendor_oid)},
        ]
    })

    orders = []
    async for po in cursor:
        po["id"]        = str(po["_id"])
        del po["_id"]
        po["vendor_id"] = str(po.get("vendor_id", ""))

        status = po.get("status", "")
        if status in ("SentToVendor", "WalkinAccepted"):
            vendor_items = po.get("vendor_response", {}).get("items")
            if vendor_items:
                po["items"] = vendor_items

        tenant_id = po.get("tenant_id")
        if tenant_id:
            tenant_key = str(tenant_id)
            if tenant_key not in tenant_names:
                tenant = await tenants_collection.find_one(
                    {"tenant_id": tenant_id},
                    {"company_name": 1, "name": 1, "tenant_id": 1},
                )
                tenant_names[tenant_key] = (
                    (tenant or {}).get("company_name")
                    or (tenant or {}).get("name")
                    or po.get("ownerSite")
                    or "Retailer"
                )
            po["retailer_name"] = tenant_names[tenant_key]
        orders.append(po)

    return orders


def _po_export_rows(po: dict) -> list:
    """Export exactly the vendor-confirmed lines when they exist."""
    items = (po.get("vendor_response") or {}).get("items") or po.get("items") or []
    rows = []
    for item in items:
        if item.get("removed"):
            continue
        quantity = item.get("amendedQty") or item.get("quantity") or 0
        rate = item.get("vendorRate") or item.get("rate") or 0
        try:
            quantity, rate = float(quantity), float(rate)
        except (TypeError, ValueError):
            quantity, rate = 0, 0
        rows.append({
            "sku": item.get("product_sku") or item.get("sku") or item.get("barcode") or "—",
            "description": item.get("description") or "—",
            "quantity": quantity,
            "rate": rate,
            "total": quantity * rate,
        })
    return rows


def _po_export_response(po: dict, vendor: dict, fmt: str) -> StreamingResponse:
    rows = _po_export_rows(po)
    total = sum(row["total"] for row in rows)
    order_no = str(po.get("orderNo") or "purchase-order").replace("/", "-").replace("\\", "-")
    retailer = po.get("retailer_name") or po.get("ownerSite") or "Retailer"
    vendor_name = vendor.get("name") or vendor.get("vendor_name") or "Vendor"
    status = po.get("status") or "Draft"
    delivery = po.get("deliveryDate") or po.get("expectedDeliveryDate") or po.get("delivery_date") or "—"

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = Workbook(); ws = wb.active; ws.title = "Purchase Order"
        ws.append(["PURCHASE ORDER", order_no])
        ws.append(["Retailer", retailer]); ws.append(["Vendor", vendor_name]); ws.append(["Status", status]); ws.append(["Delivery date", str(delivery)])
        ws.append([]); ws.append(["#", "SKU / Barcode", "Description", "Quantity", "Rate (INR)", "Line total (INR)"])
        for cell in ws[7]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="4F46E5"); cell.alignment = Alignment(horizontal="center")
        for index, row in enumerate(rows, 1):
            ws.append([index, row["sku"], row["description"], row["quantity"], row["rate"], row["total"]])
        ws.append([]); ws.append(["", "", "", "", "Grand Total", total]); ws.cell(ws.max_row, 5).font = Font(bold=True); ws.cell(ws.max_row, 6).font = Font(bold=True)
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 3, 12), 48)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{order_no}.xlsx"'})

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    buf = BytesIO(); doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15 * mm, rightMargin=15 * mm, topMargin=16 * mm, bottomMargin=16 * mm)
    styles = getSampleStyleSheet(); title = ParagraphStyle("POExportTitle", parent=styles["Heading1"], textColor=colors.HexColor("#4F46E5"))
    elements = [Paragraph("Purchase Order", title), Paragraph(f"PO: {order_no}<br/>Retailer: {retailer}<br/>Vendor: {vendor_name}<br/>Status: {status} &nbsp; | &nbsp; Delivery: {delivery}", styles["Normal"]), Spacer(1, 8 * mm)]
    table_data = [["#", "SKU / Barcode", "Description", "Qty", "Rate", "Total"]]
    for index, row in enumerate(rows, 1): table_data.append([str(index), Paragraph(escape(str(row["sku"])), styles["BodyText"]), Paragraph(escape(str(row["description"])), styles["BodyText"]), str(row["quantity"]), f"{row['rate']:,.2f}", f"{row['total']:,.2f}"])
    table_data.append(["", "", "", "", "Grand Total", f"{total:,.2f}"])
    table = Table(table_data, repeatRows=1, colWidths=[12 * mm, 42 * mm, 110 * mm, 20 * mm, 32 * mm, 35 * mm])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#CBD5E1")), ("ALIGN", (3, 1), (-1, -1), "RIGHT"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    elements.extend([table, Spacer(1, 6 * mm), Paragraph(f"Generated from RMS on {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}. This document reflects the vendor-submitted lines when available.", styles["Normal"])])
    doc.build(elements); buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{order_no}.pdf"'})


@vendor_bp.put("/purchaseorders/{po_id}/delivery")
async def update_vendor_delivery(po_id: str, payload: dict, authorization: str = Header(None)):
    """Vendor-owned dispatch data. Retailer receipt data remains read-only here."""
    vendor = await require_vendor_identity(authorization)
    query = vendor_po_query(po_id, vendor["_id"])
    po = await purchaseorders_collection.find_one(query)
    if not po:
        raise HTTPException(status_code=404, detail="PO not found or not assigned to this vendor")
    if po.get("status") not in {"VendorSubmitted", "Approved"}:
        raise HTTPException(status_code=400, detail="Submit the PO before updating delivery or dispatch.")
    status = str(payload.get("status") or "").strip()
    allowed = {"ProductionStarted", "ReadyToDispatch", "Dispatched"}
    if status not in allowed:
        raise HTTPException(status_code=400, detail="Choose Production started, Ready to dispatch, or Dispatched.")
    if status == "Dispatched" and not str(payload.get("expected_delivery_date") or "").strip():
        raise HTTPException(status_code=400, detail="Expected delivery date is required when marking a PO as dispatched.")
    delivery = dict(po.get("delivery") or {})
    vendor_update = {
        "status": status,
        "expected_dispatch_date": str(payload.get("expected_dispatch_date") or "").strip()[:30],
        "expected_delivery_date": str(payload.get("expected_delivery_date") or "").strip()[:30],
        "transporter_name": str(payload.get("transporter_name") or "").strip()[:160],
        "tracking_number": str(payload.get("tracking_number") or "").strip()[:160],
        "vehicle_number": str(payload.get("vehicle_number") or "").strip()[:80],
        "dispatch_note": str(payload.get("dispatch_note") or "").strip()[:1000],
        "updated_at": datetime.utcnow(),
        "updated_by": str(vendor.get("_id")),
    }
    delivery["vendor"] = vendor_update
    timeline = list(delivery.get("timeline") or [])
    timeline.append({"event": status, "actor": "Vendor", "at": datetime.utcnow(), "note": vendor_update["dispatch_note"]})
    delivery["timeline"] = timeline[-30:]
    await purchaseorders_collection.update_one(query, {"$set": {"delivery": delivery, "updatedAt": datetime.utcnow()}})
    if status == "Dispatched":
        from .purchaseorder_routes import consume_direct_catalogue_reservations
        from .vendor_inventory_routes import consume_reservations_for_po
        await consume_direct_catalogue_reservations(po)
        await consume_reservations_for_po(vendor["_id"], po)
    return {"message": "Delivery update saved", "delivery": delivery}

@vendor_bp.get("/purchaseorders/{po_id}/download")
async def download_vendor_purchase_order(po_id: str, format: str = Query("pdf"), authorization: str = Header(None)):
    """Download a vendor-authorized PO in PDF or Excel format."""
    vendor = await require_vendor_identity(authorization)
    po = await purchaseorders_collection.find_one(vendor_po_query(po_id, vendor["_id"]))
    if not po:
        raise HTTPException(status_code=404, detail="PO not found or not assigned to this vendor")
    fmt = format.lower()
    if fmt not in {"pdf", "xlsx"}:
        raise HTTPException(status_code=400, detail="format must be pdf or xlsx")
    return _po_export_response(po, vendor, fmt)


@vendor_bp.post("/purchaseorders/{po_id}/payment-proof/verify")
async def verify_payment_proof(po_id: str, payload: dict, authorization: str = Header(None)):
    """Vendor confirms or rejects the buyer's uploaded offline-payment proof."""
    vendor = await require_vendor_identity(authorization)
    po_query = vendor_po_query(po_id, vendor["_id"])
    po = await purchaseorders_collection.find_one(po_query)
    if not po:
        raise HTTPException(status_code=404, detail="PO not found or not assigned to this vendor")
    if not po.get("payment_proof"):
        raise HTTPException(status_code=400, detail="No payment proof has been submitted for this order yet.")

    new_status = str(payload.get("status") or "").strip().lower()
    if new_status not in ("verified", "rejected"):
        raise HTTPException(status_code=400, detail="Status must be 'verified' or 'rejected'.")

    await purchaseorders_collection.update_one(
        po_query,
        {"$set": {
            "payment_proof.status":      new_status,
            "payment_proof.verified_at": datetime.utcnow(),
            "payment_proof.verified_by": vendor.get("name") or vendor.get("vendor_name") or "Vendor",
            "payment_proof.verify_note": str(payload.get("note") or "").strip()[:500],
            "updatedAt": datetime.utcnow(),
        }}
    )
    return {"message": f"Payment marked as {new_status}."}


@vendor_bp.get("/purchaseorders/{vendor_name}")
async def get_vendor_purchase_orders(vendor_name: str):
    """Retired name-based route; vendor identity must come from the token."""
    raise HTTPException(
        status_code=410,
        detail="This route is retired. Use /api/vendors/my-purchaseorders with a vendor token.",
    )

@vendor_bp.post("/purchaseorders/{po_id}/items")
async def vendor_add_items(po_id: str, payload: dict, authorization: str = Header(None)):
    """Add or update items only on a PO assigned to the authenticated vendor."""
    vendor = await require_vendor_identity(authorization)
    po_query = vendor_po_query(po_id, vendor["_id"])
    po = await purchaseorders_collection.find_one(po_query)
    if not po:
        raise HTTPException(status_code=404, detail="PO not found or not assigned to this vendor")
    if po.get("status") not in ["SentToVendor", "VendorSubmitted", "WalkinAccepted"]:
        raise HTTPException(
            status_code=400,
            detail=f"PO not open for vendor edits. Current status: '{po.get('status')}'"
        )

    vendor_items = payload.get("items", [])
    if not isinstance(vendor_items, list) or not vendor_items:
        raise HTTPException(status_code=400, detail="No items provided")

    vendor_section = {
        "submittedAt": datetime.utcnow(),
        "items":       vendor_items,
        "status":      "Draft",
    }

    await purchaseorders_collection.update_one(
        po_query,
        {"$set": {"vendor_response": vendor_section, "updatedAt": datetime.utcnow()}}
    )

    return {"message": "Vendor items saved as draft"}


@vendor_bp.post("/purchaseorders/{po_id}/submit")
async def vendor_submit_po(po_id: str, authorization: str = Header(None)):
    """Submit only a PO assigned to the authenticated vendor."""
    vendor = await require_vendor_identity(authorization)
    po_query = vendor_po_query(po_id, vendor["_id"])
    from .purchaseorder_routes import (
        generate_item_barcode,
        calculate_po_totals,
    )

    po = await purchaseorders_collection.find_one(po_query)
    if not po:
        raise HTTPException(status_code=404, detail="PO not found or not assigned to this vendor")
    if po.get("status") not in ["SentToVendor", "VendorSubmitted", "WalkinAccepted"]:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Cannot submit: PO status is '{po.get('status')}'. "
                f"Expected SentToVendor, VendorSubmitted or WalkinAccepted."
            )
        )

    vendor_items = po.get("vendor_response", {}).get("items", [])
    if not vendor_items:
        raise HTTPException(status_code=400, detail="Vendor has not added any items yet")

    if po.get("direct_purchase"):
        expected_lines = {(str(line.get("catalogue_item_id") or ""), str(line.get("size") or ""), str(line.get("color") or "")): line for line in po.get("items") or []}
        submitted_lines = {(str(line.get("catalogue_item_id") or ""), str(line.get("size") or ""), str(line.get("color") or "")): line for line in vendor_items}
        if set(expected_lines) != set(submitted_lines):
            raise HTTPException(status_code=400, detail="A direct catalogue PO must keep the buyer's published product variants unchanged.")
        for key, expected in expected_lines.items():
            submitted = submitted_lines[key]
            if float(submitted.get("quantity") or 0) != float(expected.get("quantity") or 0) or float(submitted.get("rate") or 0) != float(expected.get("rate") or 0):
                raise HTTPException(status_code=400, detail="A direct catalogue PO has a fixed quantity and price. Reject it instead of changing its commercial lines.")
    # A vendor's barcode is a supplier reference, not RMS stock identity.
    # Preserve it separately and leave RMS barcode creation to GRC receipt.
    for item in vendor_items:
        item["vendorBarcode"] = (
            item.get("vendorBarcode") or item.get("barcode") or ""
        ).strip()

    merged_items = list(po.get("items", []))

    bc_index: dict = {}
    for i, it in enumerate(merged_items):
        bc = (it.get("barcode") or "").strip()
        if bc:
            bc_index[bc] = i

    desc_index: dict = {}
    for i, it in enumerate(merged_items):
        desc = (it.get("description") or "").strip().lower()
        bc   = (it.get("barcode") or "").strip()
        # Vendor catalogue barcodes and RMS barcodes differ, so description
        # is the safe fallback when the buyer selected an existing RMS item.
        if desc:
            desc_index.setdefault(desc, i)

    for item in vendor_items:
        vendor_bc = (item.get("vendorBarcode") or item.get("barcode") or "").strip()
        desc = (item.get("description") or "").strip().lower()
        # A matching RMS barcode is still allowed for an already-known item.
        rms_bc = (item.get("barcode") or "").strip()

        if rms_bc and rms_bc in bc_index:
            idx = bc_index[rms_bc]
            merged_items[idx] = {
                **merged_items[idx],
                "vendorBarcode": vendor_bc or merged_items[idx].get("vendorBarcode", ""),
                "amendedQty": item.get("quantity", merged_items[idx].get("amendedQty")),
                "rate":       item.get("rate",     merged_items[idx].get("rate")),
                "remarks":    item.get("remarks",  merged_items[idx].get("remarks", "")),
            }

        elif desc and desc in desc_index:
            idx = desc_index[desc]
            # Keep the PO's ITEM/... placeholder. GRC will replace it with a
            # tenant-safe RMS barcode and retain this vendor barcode separately.
            merged_items[idx] = {
                **merged_items[idx],
                "vendorBarcode": vendor_bc or merged_items[idx].get("vendorBarcode", ""),
                "amendedQty": item.get("quantity", merged_items[idx].get("amendedQty")),
                "rate":       item.get("rate",     merged_items[idx].get("rate")),
                "remarks":    item.get("remarks",  merged_items[idx].get("remarks", "")),
                "removed":    False,
            }
            del desc_index[desc]

        else:
            # New vendor-added PO lines receive only an ITEM placeholder here.
            # Their supplier label remains vendorBarcode; GRC creates RMS barcode.
            rms_placeholder = await generate_item_barcode()
            merged_items.append({
                **item,
                "barcode": rms_placeholder,
                "vendorBarcode": vendor_bc,
            })
            bc_index[rms_placeholder] = len(merged_items) - 1

    real_barcodes_desc = {
        (it.get("description") or "").strip().lower()
        for it in merged_items
        if not (it.get("barcode") or "").startswith("ITEM/")
    }
    merged_items = [
        it for it in merged_items
        if not (
            (it.get("barcode") or "").startswith("ITEM/") and
            (it.get("description") or "").strip().lower() in real_barcodes_desc
        )
    ]

    po_dict = dict(po)
    po_dict["items"] = merged_items
    calculate_po_totals(po_dict)

    await purchaseorders_collection.update_one(
        po_query,
        {"$set": {
            "status":                      "VendorSubmitted",
            "items":                       merged_items,
            "netAmount":                   po_dict["netAmount"],
            "basicValue":                  po_dict["basicValue"],
            "taxAmount":                   po_dict["taxAmount"],
            "grossAmount":                 po_dict["grossAmount"],
            "vendor_response.status":      "Submitted",
            "vendor_response.submittedAt": datetime.utcnow().isoformat(),
            "updatedAt":                   datetime.utcnow(),
        }}
    )

    return {"message": "PO submitted successfully. M-Buyer will review your submission."}


ALLOWED_CARD_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/heic", "image/heif"}
MAX_CARD_IMAGE_BYTES = 8 * 1024 * 1024  # 8MB


@vendor_bp.post("/scan-visiting-card")
async def scan_visiting_card(files: List[UploadFile] = File(...), ctx: dict = Depends(get_hq_tenant)):
    """Pre-fill the "Add Vendor from Visiting Card" form from a photo (or
    two — front and back of the same card) via a single vision-LLM call
    (see vision_extract.py) — no OCR service, no trained model. Always
    returns 200 with best-effort (possibly all-empty) fields; a scan
    failure must never block the buyer from just typing the card in
    manually."""
    if len(files) > 2:
        raise HTTPException(status_code=400, detail="Scan at most 2 images at a time (front and back of the card).")

    images: List[tuple] = []
    for file in files:
        if file.content_type not in ALLOWED_CARD_IMAGE_TYPES:
            raise HTTPException(status_code=400, detail="Please upload JPEG, PNG, WEBP or HEIC images only.")
        image_bytes = await file.read()
        if len(image_bytes) > MAX_CARD_IMAGE_BYTES:
            raise HTTPException(status_code=400, detail="Each image must be under 8MB.")
        if not image_bytes:
            raise HTTPException(status_code=400, detail="One of the uploaded files is empty.")
        images.append((image_bytes, file.content_type))

    extracted = await extract_visiting_card(images)
    return extracted


@vendor_bp.post("/invite")
async def create_vendor_invite(request: Request, ctx: dict = Depends(get_hq_tenant)):
    """HQ creates an invite link for a new vendor. Unaffected by the identity/link split — an invite has no vendor identity yet; tenant_id lives on the invite itself, exactly as before."""
    import secrets as _secrets
    body = await request.json()

    company_name     = (body.get("company_name") or body.get("companyName", "")).strip()
    brand_names      = _parse_brand_names(body.get("brand_names") or body.get("brandNames") or body.get("brand_name") or body.get("brandName"))
    contact_name     = (body.get("contact_person") or body.get("contactName", "")).strip()
    mobile           = body.get("mobile", "").strip()
    email            = (body.get("email") or "").strip()
    address          = (body.get("address") or "").strip()
    product_category = (body.get("product_type") or body.get("productCategory", "")).strip()
    expires_in_days  = int(body.get("expiresInDays", 7))

    if not company_name or not mobile:
        raise HTTPException(status_code=400, detail="company_name and mobile are required.")

    raw_token = _secrets.token_urlsafe(24)

    invite_doc = {
        "token":           raw_token,
        "companyName":     company_name,
        "brandNames":      brand_names,
        "brandName":       ", ".join(brand_names),  # kept in sync — every existing display fallback (.get("brandName")) elsewhere in the app keeps working, now showing all brands instead of just one.
        "contactName":     contact_name,
        "mobile":          mobile,
        "email":           email,
        "address":         address,
        "productCategory": product_category,
        "status":          "Pending",
        "created_by":      ctx.get("admin_id", "M-Buyer"),
        "tenant_id":       ctx["tenant_id"],
        "created_at":      datetime.utcnow(),
        "expires_at":      datetime.utcnow() + timedelta(days=expires_in_days),
        "vendor_id":       None,
    }

    await vendor_invites_collection.insert_one(invite_doc)
    return {
        "message":    "Invite created successfully",
        "token":      raw_token,
        "expires_at": invite_doc["expires_at"].isoformat(),
    }


MAX_BULK_INVITE_ROWS = 200


@vendor_bp.post("/invite/bulk")
async def create_vendor_invites_bulk(request: Request, ctx: dict = Depends(get_hq_tenant)):
    """CSV-driven version of POST /invite — same invite doc, same unique
    token per row, just looped. Unlike the single-invite flow (which lets
    the frontend decide separately whether to also call /send-invite-email),
    this sends the email immediately per row when one is provided, since
    there's no per-row UI step to defer it to."""
    import secrets as _secrets
    body = await request.json()
    rows = body.get("rows") or []
    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="No rows provided.")
    if len(rows) > MAX_BULK_INVITE_ROWS:
        raise HTTPException(status_code=400, detail=f"Bulk invite is limited to {MAX_BULK_INVITE_ROWS} rows at a time.")

    now = datetime.utcnow()
    results = []
    for row in rows:
        company_name = str(row.get("company_name") or row.get("companyName") or "").strip()
        mobile       = str(row.get("mobile") or "").strip()
        email        = str(row.get("email") or "").strip()
        contact_name = str(row.get("contact_person") or row.get("contactName") or "").strip()
        address      = str(row.get("address") or "").strip()
        product_category = str(row.get("product_type") or row.get("productCategory") or "").strip()
        brand_names  = _parse_brand_names(row.get("brand_names") or row.get("brandNames") or row.get("brand_name") or row.get("brandName"))

        if not company_name or not mobile:
            results.append({"company_name": company_name or "(blank)", "status": "error", "reason": "company_name and mobile are required."})
            continue

        raw_token = _secrets.token_urlsafe(24)
        invite_doc = {
            "token":           raw_token,
            "companyName":     company_name,
            "brandNames":      brand_names,
            "brandName":       ", ".join(brand_names),
            "contactName":     contact_name,
            "mobile":          mobile,
            "email":           email,
            "address":         address,
            "productCategory": product_category,
            "status":          "Pending",
            "created_by":      ctx.get("admin_id", "M-Buyer"),
            "tenant_id":       ctx["tenant_id"],
            "created_at":      now,
            "expires_at":      now + timedelta(days=7),
            "vendor_id":       None,
            "source":          "bulk_import",
        }
        await vendor_invites_collection.insert_one(invite_doc)

        emailed = False
        if email:
            invite_link = frontend_url(f"vendor/register?token={raw_token}")
            try:
                await send_vendor_invite_email(email, contact_name or company_name, company_name, invite_link)
                emailed = True
            except Exception:
                pass

        results.append({"company_name": company_name, "status": "created", "token": raw_token, "emailed": emailed})

    created = sum(1 for r in results if r["status"] == "created")
    return {
        "message": f"{created} of {len(rows)} invite(s) created.",
        "created_count": created,
        "results": results,
    }


@vendor_bp.get("/register-by-token")
async def get_invite_by_token(token: str):
    """Called by the vendor registration page on load. Public, unaffected by the identity/link split."""
    invite = await vendor_invites_collection.find_one({"token": token})
    if not invite:
        raise HTTPException(status_code=404, detail="Invite link is invalid or has expired.")

    if invite.get("status") == "Registered":
        raise HTTPException(status_code=400, detail="This invite link has already been used.")

    if invite.get("status") == "Expired" or datetime.utcnow() > invite["expires_at"]:
        await vendor_invites_collection.update_one(
            {"token": token}, {"$set": {"status": "Expired"}}
        )
        raise HTTPException(status_code=400, detail="This invite link has expired.")

    return {
        "companyName":     invite["companyName"],
        "brandName":       invite.get("brandName", ""),
        "brandNames":      invite.get("brandNames") or _parse_brand_names(invite.get("brandName")),
        "contactName":     invite["contactName"],
        "mobile":          invite["mobile"],
        "email":           invite.get("email", ""),
        "address":         invite.get("address", ""),
        "productCategory": invite.get("productCategory", ""),
        "expiresAt":       invite["expires_at"].isoformat(),
    }


@vendor_bp.post("/register-by-token")
async def complete_invite_registration(request: Request):
    """
    Mark an invite token as Registered and link it to the new vendor_id.
    Public — unaffected structurally; still marks the invite used and tags
    the resulting vendor. "source": "invite_link" already lives on the
    LINK document (set during /register), not here.
    """
    body      = await request.json()
    token     = body.get("token", "").strip()
    vendor_id = body.get("vendor_id", "").strip()

    if not token or not vendor_id:
        raise HTTPException(status_code=400, detail="token and vendor_id are required.")

    invite = await _require_active_vendor_invite(token)
    if not ObjectId.is_valid(vendor_id):
        raise HTTPException(status_code=400, detail="Invalid vendor registration.")

    vendor_oid = ObjectId(vendor_id)
    link = await vendor_tenant_links_collection.find_one({
        "vendor_id": vendor_oid,
        "tenant_id": invite.get("tenant_id"),
        "source": "invite_link",
    })
    if not link:
        raise HTTPException(status_code=403, detail="This vendor registration does not belong to the invitation.")

    claim = await vendor_invites_collection.update_one(
        {"_id": invite["_id"], "status": "Pending", "expires_at": {"$gt": datetime.utcnow()}},
        {"$set": {
            "status": "Registered", "vendor_id": vendor_id,
            "registered_at": datetime.utcnow(),
        }},
    )
    if claim.modified_count != 1:
        raise HTTPException(status_code=400, detail="This invite link has already been used or expired.")

    return {"message": "Invite marked as registered."}


# ── Questionnaire (public lead capture, pre-vendor) — unaffected structurally ─

# Questionnaire submissions are public, but management is tenant-scoped.
async def questionnaire_tenant_query(ctx: dict) -> dict:
    tenant = await tenants_collection.find_one(
        {"tenant_id": ctx["tenant_id"]}, {"account_type": 1}
    )
    if (tenant or {}).get("account_type") == "single_store":
        return {"tenant_id": "__single_store_questionnaires_disabled__"}
    return {"tenant_id": ctx["tenant_id"]}

@vendor_bp.post("/questionnaire")
async def submit_questionnaire(
    vendorName:          str              = Form(...),
    contactPerson:       str              = Form(""),
    phoneNumber:         str              = Form(""),
    cityLocation:        str              = Form(""),
    businessType:        str              = Form(""),
    productCategory:     str              = Form(""),
    vendorQuality:       int              = Form(0),
    moq:                 str              = Form(""),
    priceRange:          str              = Form(""),
    leadTime:            str              = Form(""),
    paymentTerms:        str              = Form(""),
    brandSection:        str              = Form(""),
    onlineCollaboration: str              = Form(""),
    email:               str              = Form(""),
    tenantId:             Optional[str]    = Form(None),
    images:              List[UploadFile] = File([]),
):
    """No auth required — public-facing lead capture form, no vendor identity created here."""
    vendorName = vendorName.strip()
    if not vendorName:
        raise HTTPException(status_code=400, detail="vendorName is required.")

    target_tenant_id = (tenantId or "").strip()
    if target_tenant_id:
        target_tenant = await tenants_collection.find_one(
            {"tenant_id": target_tenant_id}, {"account_type": 1}
        )
        if not target_tenant or target_tenant.get("account_type") == "single_store":
            raise HTTPException(status_code=400, detail="Selected retailer is not available.")

    image_urls: List[str] = []
    for img in images:
        if not img or not img.filename:
            continue
        try:
            raw    = await img.read()
            result = cloudinary.uploader.upload(
                raw,
                folder="vendor_questionnaire",
                resource_type="auto",
                public_id=f"{phoneNumber}_{img.filename}".replace(" ", "_"),
                overwrite=True,
            )
            image_urls.append(result["secure_url"])
        except Exception as e:
            print(f"⚠️ Cloudinary upload failed for {img.filename}: {e}")

    submission = {
        "vendorName":          vendorName,
        "contactPerson":       contactPerson.strip(),
        "phoneNumber":         phoneNumber.strip(),
        "email":               email.strip().lower(),
        "cityLocation":        cityLocation.strip(),
        "businessType":        businessType.strip(),
        "productCategory":     productCategory.strip(),
        "vendorQuality":       vendorQuality,
        "moq":                 moq.strip(),
        "priceRange":          priceRange.strip(),
        "leadTime":            leadTime.strip(),
        "paymentTerms":        paymentTerms.strip(),
        "brandSection":        brandSection.strip(),
        "onlineCollaboration": onlineCollaboration.strip(),
        "images":              image_urls,
        "images_count":        len(image_urls),
        "read":                False,
        "status":              "Pending",
        "submittedAt":         datetime.utcnow(),
        "invite_token":        None,
        "tenant_id":           target_tenant_id or None,
        # ⚠️ Same known gap as before this change — this public form still
        # has no tenant identifier. Not something the identity/link split
        # fixes on its own; it needs the tenant-scoped-link design flagged
        # earlier in this conversation.
    }

    result = await questionnaire_collection.insert_one(submission)

    vendor_email = email.strip().lower()
    if vendor_email:
        try:
            await send_questionnaire_received_email(
                vendor_email, vendorName, contactPerson.strip() or vendorName,
            )
        except Exception as e:
            print(f"⚠️ Questionnaire ack email failed for {vendor_email}: {e}")

    return {
        "message":       "Questionnaire submitted successfully. We will be in touch soon.",
        "submission_id": str(result.inserted_id),
        "images_saved":  len(image_urls),
    }


@vendor_bp.get("/questionnaire-submissions")
async def get_questionnaire_submissions(ctx: dict = Depends(get_hq_tenant)):
    """Unaffected by this change — see known gap noted in submit_questionnaire above."""
    scope = await questionnaire_tenant_query(ctx)
    submissions = await questionnaire_collection.find(
        scope, sort=[("submittedAt", -1)]
    ).to_list(100)
    return [serialize_doc(s) for s in submissions]


@vendor_bp.patch("/questionnaire-submissions/{submission_id}/read")
async def mark_submission_read(submission_id: str, ctx: dict = Depends(get_hq_tenant)):
    if not ObjectId.is_valid(submission_id):
        raise HTTPException(status_code=400, detail="Invalid submission ID.")
    scope = await questionnaire_tenant_query(ctx)
    result = await questionnaire_collection.update_one(
        {"_id": ObjectId(submission_id), **scope}, {"$set": {"read": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Submission not found.")
    return {"message": "Marked as read."}


@vendor_bp.post("/questionnaire-submissions/{submission_id}/accept")
async def accept_questionnaire_submission(
    submission_id: str,
    request: Request,
    ctx: dict = Depends(get_hq_tenant),
):
    """Generates an invite (tenant-scoped) from an accepted questionnaire. Unaffected structurally — still produces an invite, not a vendor identity directly."""
    if not ObjectId.is_valid(submission_id):
        raise HTTPException(status_code=400, detail="Invalid submission ID.")

    scope = await questionnaire_tenant_query(ctx)
    submission = await questionnaire_collection.find_one({"_id": ObjectId(submission_id), **scope})
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    if submission.get("status") == "Accepted":
        existing_invite = await vendor_invites_collection.find_one({"submission_id": submission_id, "tenant_id": ctx["tenant_id"]})
        if existing_invite:
            return {
                "message":   "Already accepted. Returning existing invite token.",
                "token":     existing_invite["token"],
                "expiresAt": existing_invite["expires_at"].isoformat(),
            }

    accepted_by = ctx.get("admin_id", "M-Buyer")

    import secrets
    raw_token = secrets.token_urlsafe(16)

    invite_doc = {
        "token":           raw_token,
        "companyName":     submission.get("vendorName", ""),
        "brandName":       submission.get("brandName", ""),
        "brandNames":      _parse_brand_names(submission.get("brandNames") or submission.get("brandName")),
        "contactName":     submission.get("contactPerson", ""),
        "mobile":          submission.get("phoneNumber", ""),
        "email":           "",
        "address":         submission.get("address", ""),
        "productCategory": submission.get("productCategory", ""),
        "status":          "Pending",
        "source":          "questionnaire",
        "submission_id":   submission_id,
        "created_by":      accepted_by,
        "tenant_id":       ctx["tenant_id"],
        "created_at":      datetime.utcnow(),
        "expires_at":      datetime.utcnow() + timedelta(days=7),
        "vendor_id":       None,
    }

    await vendor_invites_collection.insert_one(invite_doc)

    await questionnaire_collection.update_one(
        {"_id": ObjectId(submission_id), **scope},
        {"$set": {
            "status": "Accepted", "read": True, "accepted_by": accepted_by,
            "accepted_at": datetime.utcnow(), "invite_token": raw_token,
        }}
    )

    return {
        "message":   "Questionnaire accepted. Invite link generated.",
        "token":     raw_token,
        "expiresAt": invite_doc["expires_at"].isoformat(),
    }


@vendor_bp.patch("/questionnaire-submissions/{submission_id}/dismiss")
async def dismiss_questionnaire_submission(submission_id: str, ctx: dict = Depends(get_hq_tenant)):
    if not ObjectId.is_valid(submission_id):
        raise HTTPException(status_code=400, detail="Invalid submission ID.")
    scope = await questionnaire_tenant_query(ctx)
    result = await questionnaire_collection.update_one(
        {"_id": ObjectId(submission_id), **scope}, {"$set": {"status": "Dismissed", "read": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Submission not found.")
    return {"message": "Submission dismissed."}


@vendor_bp.get("/invites")
async def list_invites(ctx: dict = Depends(get_hq_tenant), limit: int = Query(200, ge=1, le=500)):
    """Track every invite this tenant has ever sent — from the "Add Vendor
    from Visiting Card" flow and questionnaire-accept — which vendors were
    invited, who's still pending, who actually completed registration.

    Status is computed fresh against expires_at rather than trusting the
    stored value: an invite only flips to "Expired" in the DB lazily, the
    next time someone actually opens the link (see get_invite_by_token) —
    one nobody ever clicked would otherwise show "Pending" forever, long
    after it's actually unusable.
    """
    now = datetime.utcnow()
    invites = await vendor_invites_collection.find(
        {"tenant_id": ctx["tenant_id"]}
    ).sort("created_at", -1).to_list(limit)

    inviter_ids = [ObjectId(inv["created_by"]) for inv in invites if inv.get("created_by") and ObjectId.is_valid(str(inv["created_by"]))]
    inviter_names: Dict[str, str] = {}
    if inviter_ids:
        async for a in admins_collection.find({"_id": {"$in": inviter_ids}}, {"name": 1, "email": 1}):
            inviter_names[str(a["_id"])] = a.get("name") or a.get("email", "")

    rows = []
    counts = {"Pending": 0, "Registered": 0, "Expired": 0}
    for inv in invites:
        inv_status = inv.get("status", "Pending")
        expires_at = inv.get("expires_at")
        if inv_status == "Pending" and isinstance(expires_at, datetime) and now > expires_at:
            inv_status = "Expired"
        counts[inv_status] = counts.get(inv_status, 0) + 1
        rows.append({
            "id":              str(inv["_id"]),
            "companyName":     inv.get("companyName", ""),
            "brandNames":      inv.get("brandNames") or _parse_brand_names(inv.get("brandName")),
            "contactName":     inv.get("contactName", ""),
            "mobile":          inv.get("mobile", ""),
            "email":           inv.get("email", ""),
            "productCategory": inv.get("productCategory", ""),
            "status":          inv_status,
            "source":          inv.get("source", "visiting_card"),
            "invitedBy":       inviter_names.get(str(inv.get("created_by")), inv.get("created_by") or ""),
            "createdAt":       inv["created_at"].isoformat() if isinstance(inv.get("created_at"), datetime) else None,
            "expiresAt":       expires_at.isoformat() if isinstance(expires_at, datetime) else None,
            "vendorId":        str(inv["vendor_id"]) if inv.get("vendor_id") else None,
            # Only still-usable invites carry their token forward — no
            # reason to hand back a link for one that's already registered
            # or expired.
            "token":           inv.get("token") if inv_status == "Pending" else None,
        })

    return {"data": rows, "counts": counts, "total": len(rows)}


@vendor_bp.post("/send-invite-email")
async def send_invite_email(
    request: Request,
    background_tasks: BackgroundTasks,
    ctx: dict = Depends(get_hq_tenant),
):
    body         = await request.json()
    email        = body.get("email", "").strip()
    contact_name = body.get("contact_name", "Vendor").strip()
    company_name = body.get("company_name", "").strip()
    invite_link  = body.get("invite_link", "").strip()

    if not email:
        raise HTTPException(status_code=400, detail="Email is required.")
    if not invite_link:
        raise HTTPException(status_code=400, detail="Invite link is required.")

    background_tasks.add_task(
        send_vendor_invite_email, email, contact_name, company_name, invite_link,
    )
    return {"message": f"Invite email sent to {email}"}



