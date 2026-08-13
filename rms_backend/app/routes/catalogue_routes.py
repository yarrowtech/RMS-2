
"""
catalogue_routes.py
=====================
NEW feature — vendor product catalogues + buyer inquiries.

REQUIRED — add to db.py:
    vendor_catalogue_collection    = db["vendor_catalogue"]
    catalogue_inquiries_collection = db["catalogue_inquiries"]

Wire in main.py:
    from .routes.catalogue_routes import router as catalogue_router
    app.include_router(catalogue_router)

─────────────────────────────────────────────────────────────────────────────
THE MODEL
─────────────────────────────────────────────────────────────────────────────
A catalogue item belongs to the vendor's IDENTITY (vendor_id), not any one
tenant — it's the vendor's own lookbook of what they can supply, and it's
visible to ANY retailer that has an Approved relationship with that vendor
(same visibility rule as get_products_by_vendor in products.py). A vendor
supplying both Citimart and Zudio shows the same catalogue to both, unless
they choose to mark specific items inactive.

An inquiry is the negotiation step BEFORE a real PO line item exists:
  1. Buyer (HQ admin) browses a vendor's catalogue, picks an item, and
     asks a question — "what's your price for size M in navy, qty 200?"
  2. Vendor sees the inquiry (across every retailer that's inquired,
     tenant name attached to each) and responds with a firm price/qty/
     availability, or asks a follow-up via vendor_note.
  3. Once Responded, the buyer can "convert" the inquiry — this doesn't
     silently create a PO (too much to guess about PO shape/other fields
     from here); it returns the resolved data
     (description/rate/size/color/qty) for the frontend to prefill into
     the existing PO creation form, and marks the inquiry Converted so it
     doesn't get reused.

This is deliberately a separate, lighter-weight system from
product_collection/purchaseorders — it's meant for the "vendor sends photos,
buyer picks and haggles" workflow that's currently happening over WhatsApp,
not a replacement for the formal registered-product PO flow.
"""

from fastapi import APIRouter, HTTPException, Depends, Header, Form, File, UploadFile, Request
from fastapi.responses import JSONResponse, StreamingResponse
from datetime import datetime, timedelta
from typing import Optional, List, Dict
from bson import ObjectId
from io import BytesIO
import re
import uuid
import json
import os
import zipfile
from urllib import error as urlerror, request as urlrequest
from urllib.parse import unquote, urlparse
import cloudinary
import cloudinary.uploader

from ..config import settings
from ..db import (
    vendor_catalogue_collection,
    catalogue_inquiries_collection,
    vendors_collection,
    vendor_tenant_links_collection,
    vendor_role_operations_collection,
    tenants_collection,
)
from .deps import get_hq_tenant
from .vendor_routes import decode_token  # reuse the same vendor JWT decoder
from .subscription_routes import get_vendor_tier  # single source of truth for tier limits
from .procurement_notification_routes import notify_buyer, notify_vendor

router = APIRouter(prefix="/api/catalogue", tags=["Vendor Catalogue"])

# Expired items are hidden first. Their catalogue media is removed only after
# this recovery period, so a vendor can still renew and restore listings.
CATALOGUE_MEDIA_RETENTION_DAYS = 30


def _vendor_catalogue_public_id(image_url: str) -> Optional[str]:
    """Return a Cloudinary public ID only for RMS vendor catalogue media."""
    try:
        parsed = urlparse(image_url)
        if "res.cloudinary.com" not in parsed.netloc:
            return None
        path = unquote(parsed.path or "")
        if "/upload/" not in path:
            return None
        remainder = path.split("/upload/", 1)[1].lstrip("/")
        parts = remainder.split("/")
        if parts and parts[0].startswith("v") and parts[0][1:].isdigit():
            parts = parts[1:]
        public_path = "/".join(parts)
        if not public_path.startswith("vendor_catalogue/"):
            return None
        return os.path.splitext(public_path)[0]
    except Exception:
        return None

# ── Export dependencies ──────────────────────────────────────────────────────
# NEW: requires two additional pip packages not otherwise used elsewhere in
# this codebase — install before deploying this route:
#     pip install reportlab openpyxl


def _str(v):
    return str(v) if v else ""


def _decode_vendor(authorization: Optional[str]) -> str:
    """Returns vendor_id string, or raises 401/403."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization token missing")
    decoded = decode_token(authorization.split(" ")[1])
    if not decoded:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    vendor_id = decoded.get("vendor_id")
    if not vendor_id:
        raise HTTPException(status_code=403, detail="This endpoint requires a vendor account.")
    return vendor_id


def _serialize_item(doc: dict) -> dict:
    doc = dict(doc)
    doc["_id"] = str(doc["_id"])
    doc["vendor_id"] = str(doc["vendor_id"])
    for f in ("created_at", "updated_at"):
        if doc.get(f):
            doc[f] = str(doc[f])
    return doc


def _serialize_inquiry(doc: dict) -> dict:
    doc = dict(doc)
    doc["_id"] = str(doc["_id"])
    doc["vendor_id"] = str(doc["vendor_id"])
    doc["catalogue_item_id"] = str(doc.get("catalogue_item_id") or "")
    for f in ("created_at", "responded_at", "converted_at"):
        if doc.get(f):
            doc[f] = str(doc[f])
    return doc


# ═══════════════════════════════════════════════════════════════════════════
# VENDOR SIDE — manage own catalogue
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/my-catalogue")
async def get_my_catalogue(authorization: str = Header(None)):
    """Vendor's own catalogue items, newest first."""
    vendor_id = _decode_vendor(authorization)
    items = []
    async for item in vendor_catalogue_collection.find(
        {"vendor_id": ObjectId(vendor_id)}
    ).sort("created_at", -1):
        items.append(_serialize_item(item))
    return {"status": "success", "data": items}


def _normalise_catalogue_variants(raw) -> list:
    """Accept a small, buyer-facing variant matrix without trusting client values."""
    if isinstance(raw, str):
        try:
            raw = json.loads(raw or "[]")
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Variants must be valid JSON.")
    if raw is None:
        return []
    if not isinstance(raw, list) or len(raw) > 100:
        raise HTTPException(status_code=400, detail="Provide between 0 and 100 variants.")

    variants = []
    for row in raw:
        if not isinstance(row, dict):
            raise HTTPException(status_code=400, detail="Each variant must be an object.")
        label = str(row.get("label") or "").strip()
        if not label:
            continue
        if len(label) > 120:
            raise HTTPException(status_code=400, detail="Variant label is too long.")
        try:
            price = max(0.0, float(row.get("price") or 0))
            moq = max(0, int(float(row.get("moq") or 0)))
            stock = max(0, int(float(row.get("stock") or 0)))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Variant price, MOQ and stock must be numbers.")
        variants.append({
            "label": label,
            "sku": str(row.get("sku") or "").strip()[:100],
            "price": price,
            "moq": moq,
            "stock": stock,
        })
    return variants


async def _ask_anthropic(instruction: str, max_tokens: int = 900) -> str:
    """Shared Anthropic client for optional vendor-assistant features."""
    if not settings.anthropic_api_key:
        raise HTTPException(status_code=503, detail="RMS AI Assistant is not configured yet.")
    request_payload = {
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": max_tokens,
        "messages": [{"role": "user", "content": instruction}],
    }
    req = urlrequest.Request(
        "https://api.anthropic.com/v1/messages",
        data=json.dumps(request_payload).encode("utf-8"),
        headers={"content-type": "application/json", "x-api-key": settings.anthropic_api_key, "anthropic-version": "2023-06-01"},
        method="POST",
    )
    try:
        with urlrequest.urlopen(req, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))["content"][0]["text"].strip()
    except (KeyError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=502, detail="RMS AI Assistant could not respond. Please try again.") from exc


@router.post("/vendor-copilot/ask")
async def ask_vendor_copilot(payload: dict, authorization: str = Header(None)):
    """Optional AI help. It never performs RMS actions or reads vendor records."""
    _decode_vendor(authorization)
    if payload.get("consent") is not True:
        raise HTTPException(status_code=400, detail="Consent is required before sending a question to AI.")
    question = str(payload.get("question") or "").strip()[:2000]
    page = str(payload.get("page") or "RMS workspace").strip()[:100]
    if not question:
        raise HTTPException(status_code=400, detail="Enter a question first.")
    instruction = (
        "You are RMS Co-pilot, a concise, friendly guide for a wholesale vendor portal. "
        "The vendor is currently on the '" + page + "' page. Answer only with safe workflow guidance. "
        "Do not claim to see their account, orders, prices, customers, documents, or data. "
        "Never tell them that you changed, submitted, published, contacted, or approved anything. "
        "For account-specific errors, direct them to RMS Help & Support. Use short practical steps. "
        "Current vendor rules: My Inventory is vendor-owned and separate from retailer and B2B stock. Buyer approval auto-reserves matching vendor SKU/barcode stock; on-hand stock deducts only when the vendor marks the retailer PO dispatched. B2B Trade is separate: RFQ, quote, B2B PO, supplier confirmation, dispatch/challan, buyer receipt, invoice and payment. B2B returns are buyer request, supplier approval, buyer return dispatch, supplier receipt, B2B stock reversal and credit note. "
        "Vendor question: " + question
    )
    return {"reply": await _ask_anthropic(instruction, max_tokens=700)}


@router.post("/my-catalogue/assistant")
async def draft_catalogue_item_with_ai(payload: dict, authorization: str = Header(None)):
    """Suggest a catalogue draft. The vendor must review it before saving."""
    _decode_vendor(authorization)
    if not settings.anthropic_api_key:
        raise HTTPException(status_code=503, detail="Catalogue Assistant is not configured yet.")

    prompt = str(payload.get("prompt") or "").strip()[:2000]
    context = {
        "item_name": str(payload.get("item_name") or "").strip()[:160],
        "category": str(payload.get("category") or "").strip()[:100],
        "available_sizes": str(payload.get("available_sizes") or "").strip()[:300],
        "available_colors": str(payload.get("available_colors") or "").strip()[:300],
    }
    if not prompt and not any(context.values()):
        raise HTTPException(status_code=400, detail="Describe the product or provide some product details first.")

    instruction = (
        "You assist a wholesale vendor creating a retailer-facing catalogue listing. "
        "Return ONLY a JSON object with keys: item_name, category, description, available_sizes, available_colors, "
        "suggested_moq, variants. variants must be an array of up to 12 objects with label, sku, price, moq, stock. "
        "Use empty strings or an empty array for information not supplied; do not invent prices, tax, certification, or stock. "
        "Write a concise factual description suitable for retail buyers. Product context: " + json.dumps(context) + ". Vendor notes: " + prompt
    )
    try:
        raw = await _ask_anthropic(instruction)
        if raw.startswith("```"):
            raw = raw.strip("`").removeprefix("json").strip()
        suggestion = json.loads(raw)
    except (KeyError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=502, detail="Catalogue Assistant could not generate a draft. Please try again.") from exc

    return {"data": {
        "item_name": str(suggestion.get("item_name") or "").strip()[:160],
        "category": str(suggestion.get("category") or "").strip()[:100],
        "description": str(suggestion.get("description") or "").strip()[:2000],
        "available_sizes": str(suggestion.get("available_sizes") or "").strip()[:300],
        "available_colors": str(suggestion.get("available_colors") or "").strip()[:300],
        "moq": max(0, int(float(suggestion.get("suggested_moq") or 0))),
        "variants": _normalise_catalogue_variants(suggestion.get("variants") or []),
    }}

@router.post("/my-catalogue", status_code=201)
async def add_catalogue_item(
    authorization:      str            = Header(None),
    item_name:           str            = Form(...),
    category:            str            = Form(""),
    description:         str            = Form(""),
    price_range_min:      float          = Form(0),
    price_range_max:      float          = Form(0),
    price:                float          = Form(0),        # firm, non-negotiated price — required for direct_purchase_enabled
    direct_purchase_enabled: bool        = Form(False),     # buyer can order this item instantly, skipping the inquiry/negotiation flow
    stock:                int            = Form(0),         # units available — required for direct_purchase_enabled, capped at order time
    available_sizes:      str            = Form(""),   # comma-separated, e.g. "S,M,L,XL"
    available_colors:     str            = Form(""),   # comma-separated
    moq:                  int            = Form(0),
    variants:             str            = Form("[]"),
    images:                List[UploadFile] = File([]),
):
    """
    Vendor adds a catalogue item — one or more images plus basic info
    (price range, available sizes/colors, MOQ). No tenant_id — visible to
    any retailer with an Approved relationship, per the module docstring.

    Gated by subscription tier (see subscription_routes.py — TIER_CONFIG is
    the single source of truth): each tier caps total active catalogue
    items, and the visibility window (expires_at) is stamped from the
    tier's visibility_days at upload time. A vendor on Free with 5 active
    items can't add a 6th until either they delete one or their items
    naturally expire — upgrading is the other way past the wall.
    """
    vendor_id = _decode_vendor(authorization)

    if not item_name.strip():
        raise HTTPException(status_code=400, detail="item_name is required.")
    if direct_purchase_enabled and price <= 0:
        raise HTTPException(status_code=400, detail="A firm price is required to enable direct purchase on this item.")
    if direct_purchase_enabled and stock <= 0:
        raise HTTPException(status_code=400, detail="Available stock is required to enable direct purchase on this item.")

    # ── Tier check BEFORE any Cloudinary upload — no point burning upload
    # calls on a request that's going to be rejected anyway.
    tier = await get_vendor_tier(vendor_id)
    current_count = await vendor_catalogue_collection.count_documents({
        "vendor_id": ObjectId(vendor_id),
        "active": True,
    })
    if current_count >= tier["image_limit"]:
        raise HTTPException(
            status_code=403,
            detail=(
                f"Your {tier['label']} plan allows {tier['image_limit']} active catalogue items. "
                f"You have {current_count}. Delete an item or upgrade your plan to add more."
            )
        )

    # ── NEW — per-item photo limit (was unlimited on every tier before
    # this pass). Checked before uploading, same reasoning as the item
    # count check above: don't burn Cloudinary calls on images that will
    # just be dropped.
    photo_limit = tier["photos_per_item"]
    real_images = [img for img in images if img and img.filename]
    if photo_limit is not None and len(real_images) > photo_limit:
        raise HTTPException(
            status_code=403,
            detail=(
                f"Your {tier['label']} plan allows {photo_limit} photo(s) per item. "
                f"You uploaded {len(real_images)}. Upgrade your plan for more photos per item."
            )
        )

    image_urls: List[str] = []
    for img in real_images:
        try:
            raw = await img.read()
            result = cloudinary.uploader.upload(
                raw,
                folder="vendor_catalogue",
                resource_type="auto",
                public_id=f"{vendor_id}_{item_name}_{img.filename}".replace(" ", "_"),
                overwrite=False,
            )
            image_urls.append(result["secure_url"])
        except Exception as e:
            print(f"⚠️ Cloudinary upload failed for {img.filename}: {e}")

    if not image_urls:
        raise HTTPException(status_code=400, detail="At least one image is required.")

    normalised_variants = _normalise_catalogue_variants(variants)

    now = datetime.utcnow()
    doc = {
        "vendor_id":         ObjectId(vendor_id),
        "item_name":         item_name.strip(),
        "category":          category.strip(),
        "description":       description.strip(),
        "images":            image_urls,
        "price_range_min":   max(0.0, price_range_min),
        "price_range_max":   max(0.0, price_range_max),
        "price":                  max(0.0, price),
        "direct_purchase_enabled": bool(direct_purchase_enabled),
        "stock":             max(0, stock),
        "available_sizes":   [s.strip() for s in available_sizes.split(",") if s.strip()],
        "available_colors":  [c.strip() for c in available_colors.split(",") if c.strip()],
        "moq":               max(0, moq),
        "variants":          normalised_variants,
        "active":            True,
        "tier_at_upload":    tier["tier"],
        "created_at":        now,
        "updated_at":        now,
        # ⚠️ NEW — the 45/90-day auto-expiry mechanism. This is only a
        # timestamp; nothing actually hides the item until the sweep job
        # (expire_stale_catalogue_items, exposed as POST
        # /api/catalogue/expire-sweep below) runs and sets active=False.
        # Must be wired to a real scheduler (cron / APScheduler / Celery
        # beat) — see that route's docstring.
        "expires_at":        now + timedelta(days=tier["visibility_days"]),
    }
    result = await vendor_catalogue_collection.insert_one(doc)
    return {
        "status": "success",
        "message": f"Catalogue item added. Visible until {doc['expires_at'].strftime('%d %b %Y')}.",
        "id": str(result.inserted_id),
        "expires_at": doc["expires_at"].isoformat(),
    }


MAX_BULK_CATALOGUE_ROWS = 100
MAX_BULK_IMAGE_FILES = 200
MAX_BULK_IMAGE_BYTES = 10 * 1024 * 1024
MAX_BULK_IMAGE_TOTAL_BYTES = 100 * 1024 * 1024


def _catalogue_image_key(value: str) -> str:
    """Stable filename-to-row matching key for bulk catalogue media."""
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _is_catalogue_image(filename: str) -> bool:
    return os.path.splitext(str(filename or "").lower())[1] in {
        ".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tiff", ".tif", ".heic", ".heif", ".avif",
    }


def _bulk_row_image_keys(row: dict) -> list:
    return [
        _catalogue_image_key(row.get("image_key")),
        _catalogue_image_key(row.get("sku")),
        _catalogue_image_key(row.get("item_name")),
    ]


async def _collect_bulk_image_uploads(images: List[UploadFile], archive: Optional[UploadFile]) -> list:
    """Read a bounded set of image files, optionally from one ZIP archive."""
    uploads = []
    total_bytes = 0
    for image in [image for image in images if image and image.filename]:
        if not _is_catalogue_image(image.filename):
            continue
        raw = await image.read()
        if not raw or len(raw) > MAX_BULK_IMAGE_BYTES or total_bytes + len(raw) > MAX_BULK_IMAGE_TOTAL_BYTES:
            continue
        total_bytes += len(raw)
        uploads.append((os.path.basename(image.filename), raw))

    if archive and archive.filename:
        if not archive.filename.lower().endswith(".zip"):
            raise HTTPException(status_code=400, detail="The image archive must be a .zip file.")
        archive_bytes = await archive.read()
        if len(archive_bytes) > 50 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="The image ZIP must be 50 MB or smaller.")
        try:
            with zipfile.ZipFile(BytesIO(archive_bytes)) as package:
                for entry in package.infolist():
                    if entry.is_dir() or not _is_catalogue_image(entry.filename) or entry.file_size > MAX_BULK_IMAGE_BYTES:
                        continue
                    raw = package.read(entry)
                    if total_bytes + len(raw) > MAX_BULK_IMAGE_TOTAL_BYTES:
                        break
                    total_bytes += len(raw)
                    uploads.append((os.path.basename(entry.filename), raw))
                    if len(uploads) >= MAX_BULK_IMAGE_FILES:
                        break
        except zipfile.BadZipFile as exc:
            raise HTTPException(status_code=400, detail="The image ZIP could not be read.") from exc

    if len(uploads) > MAX_BULK_IMAGE_FILES:
        raise HTTPException(status_code=400, detail=f"Upload at most {MAX_BULK_IMAGE_FILES} product images at a time.")
    return uploads

def _match_bulk_images_to_row(row: dict, uploads: list) -> list:
    """Match SKU/image_key/product-name to exact or numbered image filenames."""
    keys = [key for key in _bulk_row_image_keys(row) if key]
    matched = []
    for filename, raw in uploads:
        stem = _catalogue_image_key(os.path.splitext(filename)[0])
        if any(stem == key or stem.startswith(f"{key}image") or stem.startswith(f"{key}photo") or stem.startswith(f"{key}img") or (stem.startswith(key) and stem[len(key):].isdigit()) for key in keys):
            matched.append((filename, raw))
    return matched


@router.post("/my-catalogue/bulk")
async def add_catalogue_items_bulk(request: Request, authorization: str = Header(None)):
    """
    Link-based bulk version of POST /my-catalogue — for a vendor who
    already has product photos hosted somewhere (their own site, Google
    Drive, etc.) instead of uploading each item's images one at a time
    through this form. Each row always becomes direct_purchase_enabled
    with a firm price — a price RANGE (like the single-item form allows)
    doesn't make sense for something a buyer can order with one click, see
    the module docstring for why direct purchase needs a firm price.

    Images are stored as the external URL(s) as-is — NOT downloaded and
    re-uploaded to Cloudinary. Broken or expired links are the vendor's
    responsibility; RMS doesn't validate or mirror them.
    """
    vendor_id = _decode_vendor(authorization)
    body = await request.json()
    rows = body.get("rows") or []
    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="No rows provided.")
    if len(rows) > MAX_BULK_CATALOGUE_ROWS:
        raise HTTPException(status_code=400, detail=f"Bulk import is limited to {MAX_BULK_CATALOGUE_ROWS} rows at a time.")

    tier = await get_vendor_tier(vendor_id)
    current_count = await vendor_catalogue_collection.count_documents({
        "vendor_id": ObjectId(vendor_id), "active": True,
    })
    photo_limit = tier["photos_per_item"]

    now = datetime.utcnow()
    results = []
    for row in rows:
        item_name = str(row.get("item_name") or "").strip()
        image_field = str(row.get("image_url") or row.get("image_urls") or "").strip()
        image_urls = [u.strip() for u in re.split(r"[;,]", image_field) if u.strip()]

        if not item_name or not image_urls:
            results.append({"item_name": item_name or "(blank)", "status": "error", "reason": "item_name and at least one image_url are required."})
            continue
        try:
            price = float(row.get("price") or 0)
        except (TypeError, ValueError):
            price = 0
        if price <= 0:
            results.append({"item_name": item_name, "status": "error", "reason": "A firm price is required for direct purchase."})
            continue
        try:
            stock = int(float(row.get("stock") or 0))
        except (TypeError, ValueError):
            stock = 0
        if stock <= 0:
            results.append({"item_name": item_name, "status": "error", "reason": "Available stock is required for direct purchase."})
            continue
        if current_count >= tier["image_limit"]:
            results.append({"item_name": item_name, "status": "error", "reason": f"Your {tier['label']} plan allows {tier['image_limit']} active catalogue items — limit reached."})
            continue
        if photo_limit is not None and len(image_urls) > photo_limit:
            image_urls = image_urls[:photo_limit]

        moq_raw = str(row.get("moq") or "").strip()
        doc = {
            "vendor_id":               ObjectId(vendor_id),
            "item_name":               item_name,
            "category":                str(row.get("category") or "").strip(),
            "description":             str(row.get("description") or "").strip(),
            "images":                  image_urls,
            "price_range_min":         price,
            "price_range_max":         price,
            "price":                   price,
            "direct_purchase_enabled": True,
            "stock":                   stock,
            "available_sizes":         [s.strip() for s in str(row.get("available_sizes") or "").split(",") if s.strip()],
            "available_colors":        [c.strip() for c in str(row.get("available_colors") or "").split(",") if c.strip()],
            "moq":                     int(moq_raw) if moq_raw.isdigit() else 0,
            "variants":                _normalise_catalogue_variants(row.get("variants") or []),
            "active":                  True,
            "tier_at_upload":          tier["tier"],
            "source":                  "bulk_import",
            "created_at":              now,
            "updated_at":              now,
            "expires_at":              now + timedelta(days=tier["visibility_days"]),
        }
        await vendor_catalogue_collection.insert_one(doc)
        current_count += 1
        results.append({"item_name": item_name, "status": "created"})

    created = sum(1 for r in results if r["status"] == "created")
    return {
        "status": "success",
        "message": f"{created} of {len(rows)} catalogue item(s) created.",
        "created_count": created,
        "results": results,
    }




@router.post("/my-catalogue/bulk-upload")
async def add_catalogue_items_bulk_with_images(
    rows_json: str = Form(...),
    authorization: str = Header(None),
    images: List[UploadFile] = File([]),
    archive: Optional[UploadFile] = File(None),
):
    """Create many direct-purchase listings from a sheet and local product images.

    Image filenames match the sheet's image_key, SKU, or item_name. For example,
    SKU-RED-01.jpg and SKU-RED-01-2.jpg attach to the SKU-RED-01 row.
    """
    vendor_id = _decode_vendor(authorization)
    try:
        rows = json.loads(rows_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Product rows could not be read.") from exc
    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="Provide at least one product row.")
    if len(rows) > MAX_BULK_CATALOGUE_ROWS:
        raise HTTPException(status_code=400, detail=f"Bulk import is limited to {MAX_BULK_CATALOGUE_ROWS} rows at a time.")

    uploads = await _collect_bulk_image_uploads(images, archive)
    if not uploads:
        raise HTTPException(status_code=400, detail="Upload at least one JPG, PNG or WEBP product image, or a ZIP containing images.")

    tier = await get_vendor_tier(vendor_id)
    current_count = await vendor_catalogue_collection.count_documents({"vendor_id": ObjectId(vendor_id), "active": True})
    photo_limit = tier["photos_per_item"]
    now = datetime.utcnow()
    results = []

    for row in rows:
        if not isinstance(row, dict):
            results.append({"item_name": "(invalid row)", "status": "error", "reason": "Each product row must be an object."})
            continue
        item_name = str(row.get("item_name") or "").strip()
        if not item_name:
            results.append({"item_name": "(blank)", "status": "error", "reason": "item_name is required."})
            continue
        try:
            price = float(row.get("price") or 0)
        except (TypeError, ValueError):
            price = 0
        if price <= 0:
            results.append({"item_name": item_name, "status": "error", "reason": "A firm price is required for direct purchase."})
            continue
        try:
            stock = int(float(row.get("stock") or 0))
        except (TypeError, ValueError):
            stock = 0
        if stock <= 0:
            results.append({"item_name": item_name, "status": "error", "reason": "Available stock is required for direct purchase."})
            continue
        if current_count >= tier["image_limit"]:
            results.append({"item_name": item_name, "status": "error", "reason": f"Your {tier['label']} plan allows {tier['image_limit']} active catalogue items — limit reached."})
            continue

        matched = _match_bulk_images_to_row(row, uploads)
        if photo_limit is not None:
            matched = matched[:photo_limit]
        if not matched:
            expected_key = str(row.get("image_key") or row.get("sku") or item_name)
            results.append({"item_name": item_name, "status": "error", "reason": f"No uploaded image matched '{expected_key}'. Rename images to the SKU, image_key or product name."})
            continue

        image_urls = []
        for filename, raw in matched:
            try:
                result = cloudinary.uploader.upload(
                    raw,
                    folder="vendor_catalogue",
                    resource_type="image",
                    public_id=f"{vendor_id}_bulk_{uuid.uuid4().hex[:12]}_{filename}".replace(" ", "_"),
                    overwrite=False,
                )
                image_urls.append(result["secure_url"])
            except Exception as exc:
                print(f"Bulk catalogue upload failed for {filename}: {exc}")
        if not image_urls:
            results.append({"item_name": item_name, "status": "error", "reason": "Images could not be uploaded. Please try again."})
            continue

        moq_raw = str(row.get("moq") or "").strip()
        vendor_catalogue = {
            "vendor_id": ObjectId(vendor_id),
            "item_name": item_name,
            "sku": str(row.get("sku") or "").strip()[:100],
            "image_key": str(row.get("image_key") or "").strip()[:120],
            "category": str(row.get("category") or "").strip(),
            "description": str(row.get("description") or "").strip(),
            "images": image_urls,
            "price_range_min": price,
            "price_range_max": price,
            "price": price,
            "direct_purchase_enabled": True,
            "stock": stock,
            "available_sizes": [value.strip() for value in str(row.get("available_sizes") or "").split(",") if value.strip()],
            "available_colors": [value.strip() for value in str(row.get("available_colors") or "").split(",") if value.strip()],
            "moq": int(moq_raw) if moq_raw.isdigit() else 0,
            "variants": _normalise_catalogue_variants(row.get("variants") or []),
            "active": True,
            "tier_at_upload": tier["tier"],
            "source": "bulk_image_upload",
            "created_at": now,
            "updated_at": now,
            "expires_at": now + timedelta(days=tier["visibility_days"]),
        }
        await vendor_catalogue_collection.insert_one(vendor_catalogue)
        current_count += 1
        results.append({"item_name": item_name, "status": "created", "matched_images": len(image_urls)})

    created = sum(1 for result in results if result["status"] == "created")
    return {
        "status": "success",
        "message": f"{created} of {len(rows)} product(s) created using your {tier['label']} plan limits.",
        "tier": {"label": tier["label"], "item_limit": tier["image_limit"], "photos_per_item": tier["photos_per_item"]},
        "created_count": created,
        "results": results,
    }

@router.patch("/my-catalogue/{item_id}")
async def update_catalogue_item(item_id: str, payload: dict, authorization: str = Header(None)):
    """Vendor edits price/sizes/colors/active status on their own item. Images not editable here — delete and re-add for image changes."""
    vendor_id = _decode_vendor(authorization)
    if not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="Invalid item ID.")

    item = await vendor_catalogue_collection.find_one({"_id": ObjectId(item_id), "vendor_id": ObjectId(vendor_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Catalogue item not found.")

    allowed = {"item_name", "category", "description", "price_range_min", "price_range_max",
               "price", "direct_purchase_enabled", "stock",
               "available_sizes", "available_colors", "moq", "variants", "active"}
    patch = {k: v for k, v in payload.items() if k in allowed}
    if patch.get("direct_purchase_enabled") and float(patch.get("price", item.get("price", 0)) or 0) <= 0:
        raise HTTPException(status_code=400, detail="A firm price is required to enable direct purchase on this item.")
    if patch.get("direct_purchase_enabled") and int(patch.get("stock", item.get("stock", 0)) or 0) <= 0:
        raise HTTPException(status_code=400, detail="Available stock is required to enable direct purchase on this item.")
    if "variants" in patch:
        patch["variants"] = _normalise_catalogue_variants(patch["variants"])
    patch["updated_at"] = datetime.utcnow()

    # Reactivating a previously-expired item ("active": False -> True) must
    # also refresh its visibility window — otherwise the next expire-sweep
    # run (POST /api/catalogue/expire-sweep) immediately hides it again
    # since expires_at is still in the past. This is also the only way a
    # Free vendor (no paid renewal to bulk-extend for them) gets an item
    # visible again once it lapses.
    update_ops = {"$set": patch}
    if patch.get("active") is True and not item.get("active", False):
        if not item.get("images"):
            raise HTTPException(status_code=400, detail="This listing's expired images were deleted after the 30-day recovery period. Add new images before reactivating it.")
        tier = await get_vendor_tier(vendor_id)
        active_count = await vendor_catalogue_collection.count_documents({
            "vendor_id": ObjectId(vendor_id), "active": True,
        })
        if active_count >= tier["image_limit"]:
            raise HTTPException(
                status_code=403,
                detail=(
                    f"Your {tier['label']} plan allows {tier['image_limit']} active catalogue items. "
                    f"Deactivate another item or upgrade your plan to reactivate this one."
                )
            )
        patch["expires_at"] = datetime.utcnow() + timedelta(days=tier["visibility_days"])
        update_ops["$unset"] = {"expired_at": "", "expired_reason": ""}

    await vendor_catalogue_collection.update_one({"_id": ObjectId(item_id)}, update_ops)
    return {"status": "success", "message": "Catalogue item updated."}


@router.post("/my-catalogue/{item_id}/images")
async def add_item_images(
    item_id:      str,
    authorization: str             = Header(None),
    images:        List[UploadFile] = File([]),
):
    """
    ⚠️ FIXED — this is the missing piece: previously the ONLY way to
    change an item's images was delete the whole item and recreate it
    from scratch (losing its creation date, expiry countdown, and forcing
    every field to be retyped). Now a vendor can add more photos to an
    existing item directly.

    Still respects the tier's per-item photo limit — counts EXISTING
    images on the item plus the new ones being added, not just the new
    batch in isolation (a vendor on a 3-photos-per-item plan with 2
    existing images can add at most 1 more, not 3 more).
    """
    vendor_id = _decode_vendor(authorization)
    if not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="Invalid item ID.")

    item = await vendor_catalogue_collection.find_one({"_id": ObjectId(item_id), "vendor_id": ObjectId(vendor_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Catalogue item not found.")

    real_images = [img for img in images if img and img.filename]
    if not real_images:
        raise HTTPException(status_code=400, detail="At least one image is required.")

    tier = await get_vendor_tier(vendor_id)
    photo_limit = tier["photos_per_item"]
    existing_count = len(item.get("images", []))
    if photo_limit is not None and (existing_count + len(real_images)) > photo_limit:
        raise HTTPException(
            status_code=403,
            detail=(
                f"Your {tier['label']} plan allows {photo_limit} photo(s) per item. "
                f"This item already has {existing_count}; you can add at most {max(0, photo_limit - existing_count)} more."
            )
        )

    new_urls: List[str] = []
    for img in real_images:
        try:
            raw = await img.read()
            result = cloudinary.uploader.upload(
                raw,
                folder="vendor_catalogue",
                resource_type="auto",
                public_id=f"{vendor_id}_{item_id}_{img.filename}_{uuid.uuid4().hex[:8]}".replace(" ", "_"),
                overwrite=False,
            )
            new_urls.append(result["secure_url"])
        except Exception as e:
            print(f"⚠️ Cloudinary upload failed for {img.filename}: {e}")

    if not new_urls:
        raise HTTPException(status_code=400, detail="All image uploads failed. Try again.")

    await vendor_catalogue_collection.update_one(
        {"_id": ObjectId(item_id)},
        {"$push": {"images": {"$each": new_urls}}, "$set": {"updated_at": datetime.utcnow()}}
    )
    return {"status": "success", "message": f"{len(new_urls)} image(s) added.", "added": new_urls}


@router.delete("/my-catalogue/{item_id}/images")
async def remove_item_image(item_id: str, payload: dict, authorization: str = Header(None)):
    """
    Removes ONE specific image (by URL) from an item. A catalogue item
    always needs at least one image — this refuses to remove the last one
    rather than leaving an item with zero photos (which would break the
    grid card and stickers).
    """
    vendor_id = _decode_vendor(authorization)
    if not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="Invalid item ID.")

    image_url = (payload.get("image_url") or "").strip()
    if not image_url:
        raise HTTPException(status_code=400, detail="image_url is required.")

    item = await vendor_catalogue_collection.find_one({"_id": ObjectId(item_id), "vendor_id": ObjectId(vendor_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Catalogue item not found.")

    if len(item.get("images", [])) <= 1:
        raise HTTPException(status_code=400, detail="Can't remove the last image — add a replacement first, or delete the whole item instead.")
    if image_url not in item.get("images", []):
        raise HTTPException(status_code=404, detail="That image isn't on this item.")

    await vendor_catalogue_collection.update_one(
        {"_id": ObjectId(item_id)},
        {"$pull": {"images": image_url}, "$set": {"updated_at": datetime.utcnow()}}
    )
    return {"status": "success", "message": "Image removed."}


@router.delete("/my-catalogue/{item_id}")
async def delete_catalogue_item(item_id: str, authorization: str = Header(None)):
    vendor_id = _decode_vendor(authorization)
    if not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="Invalid item ID.")

    result = await vendor_catalogue_collection.delete_one({"_id": ObjectId(item_id), "vendor_id": ObjectId(vendor_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Catalogue item not found.")
    return {"status": "success", "message": "Catalogue item deleted."}


# ═══════════════════════════════════════════════════════════════════════════
# BUYER (HQ) SIDE — browse a specific vendor's catalogue
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/vendor/{vendor_id}/storefront")
async def get_vendor_storefront(vendor_id: str, ctx: dict = Depends(get_hq_tenant)):
    """Buyer-safe vendor storefront, visible only to an approved retailer relationship."""
    if not ObjectId.is_valid(vendor_id):
        raise HTTPException(status_code=400, detail="Invalid vendor ID.")
    vendor_oid = ObjectId(vendor_id)
    link = await vendor_tenant_links_collection.find_one({"vendor_id": vendor_oid, "tenant_id": ctx["tenant_id"], "status": "Approved"})
    if not link:
        raise HTTPException(status_code=404, detail="Vendor not found or not approved for this retailer.")
    vendor = await vendors_collection.find_one({"_id": vendor_oid})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found.")
    address = vendor.get("address")
    location = address if isinstance(address, str) else ", ".join(str(address.get(key) or "").strip() for key in ("city", "state", "country") if address.get(key)) if isinstance(address, dict) else ""
    items = []
    async for item in vendor_catalogue_collection.find({"vendor_id": vendor_oid, "active": True}).sort("created_at", -1):
        items.append(_serialize_item(item))
    tier = await get_vendor_tier(vendor_id)
    return {"status": "success", "data": {
        "vendor": {
            "_id": str(vendor_oid),
            "name": vendor.get("brandName") or vendor.get("name") or vendor.get("vendor_name") or "Vendor",
            "contact_name": vendor.get("contact_person") or vendor.get("owner_name") or "",
            "phone": str(vendor.get("phone") or vendor.get("mobile") or ""),
            "email": vendor.get("email") or "",
            "website": vendor.get("website") or vendor.get("website_url") or "",
            "instagram": vendor.get("instagram") or vendor.get("instagram_url") or "",
            "location": location or vendor.get("city") or "",
            "business_type": vendor.get("business_type") or [],
            "verified": link.get("kyb_status") == "Verified",
            "tier": tier["label"],
        },
        "items": items,
    }}

@router.get("/vendor/{vendor_id}")
async def get_vendor_catalogue(vendor_id: str, ctx: dict = Depends(get_hq_tenant)):
    """
    Buyer views one vendor's catalogue — only if that vendor has an
    Approved relationship with the caller's tenant. Same visibility check
    as get_products_by_vendor in products.py.
    """
    try:
        vendor_oid = ObjectId(vendor_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid vendor ID.")

    approved_link = await vendor_tenant_links_collection.find_one({
        "vendor_id": vendor_oid,
        "tenant_id": ctx["tenant_id"],
        "status":    "Approved",
    })
    if not approved_link:
        raise HTTPException(
            status_code=404,
            detail="Vendor not found or does not have an approved relationship with your retailer."
        )

    items = []
    async for item in vendor_catalogue_collection.find(
        {"vendor_id": vendor_oid, "active": True}
    ).sort("created_at", -1):
        items.append(_serialize_item(item))

    return {"status": "success", "count": len(items), "data": items}


@router.get("/search")
async def search_vendor_catalogues(
    category:      Optional[str] = None,
    business_type: Optional[str] = None,
    vendor_name:   Optional[str] = None,
    material:      Optional[str] = None,
    audience:      Optional[str] = None,
    capability:    Optional[str] = None,
    service_region: Optional[str] = None,
    min_price:     Optional[float] = None,
    max_price:     Optional[float] = None,
    ctx: dict = Depends(get_hq_tenant),
):
    """
    Buyer discovery across ALL of their tenant's Approved vendors at once —
    the missing piece that makes the multi-vendor comparison flow
    (POST/GET /inquiries/compare) actually reachable. Without this, a
    buyer could only ever browse one vendor they already knew by ID.

    Deliberately scoped to Approved vendors only — same boundary as every
    other buyer-facing route in this file (get_vendor_catalogue,
    create_inquiry). This does NOT let a buyer discover vendors they have
    no relationship with; it only lets them filter/search AMONG vendors
    already approved for their tenant. Widening that (e.g. a public
    vendor marketplace search) is a real product/policy decision, not
    something to fold in here.

    Query params (both optional, combine with AND if both given):
      category      — substring match against vendor_catalogue_collection
                       items' category/item_name/description
      business_type — exact match against the vendor identity's
                       business_type array (general_vendor/wholesaler/
                       manufacturer/retailer/fabric_supplier/exporter)

    Returns one row per matching catalogue item (not per vendor) so the
    buyer sees individual products to select for a comparison — same shape
    _serialize_item already produces, with vendor_name/vendor_id/
    business_type attached.
    """
    approved_vendor_ids = []
    async for link in vendor_tenant_links_collection.find({
        "tenant_id": ctx["tenant_id"], "status": "Approved",
    }, {"vendor_id": 1}):
        approved_vendor_ids.append(link["vendor_id"])

    if not approved_vendor_ids:
        return {"status": "success", "count": 0, "data": []}

    vendor_query: dict = {"_id": {"$in": approved_vendor_ids}}
    if business_type:
        vendor_query["business_type"] = business_type
    if vendor_name:
        vendor_query["$or"] = [
            {"name": {"$regex": vendor_name.strip(), "$options": "i"}},
            {"vendor_name": {"$regex": vendor_name.strip(), "$options": "i"}},
        ]

    vendor_map: dict = {}
    async for v in vendors_collection.find(vendor_query):
        # NEW — priority_rank and featured_badge, so results can be sorted
        # by tier and a "Verified" badge shown for Premium vendors. Real
        # differentiators for paying, not just "more storage."
        vendor_tier = await get_vendor_tier(str(v["_id"]))
        vendor_map[v["_id"]] = {
            "vendor_name":    v.get("name") or v.get("vendor_name", ""),
            "business_type":  v.get("business_type", []),
            "priority_rank":  vendor_tier["priority_rank"],
            "featured_badge": vendor_tier["featured_badge"],
        }

    if not vendor_map:
        return {"status": "success", "count": 0, "data": []}

    # Specialist profiles are separate from catalogue items. They filter
    # discovery but deliberately keep the existing catalogue → inquiry → PO
    # response shape and flow unchanged.
    profile_query: dict = {"vendor_id": {"$in": list(vendor_map.keys())}}
    profile_filters = []
    if capability and capability.strip():
        capability_pattern = {"$regex": capability.strip(), "$options": "i"}
        profile_filters.append({"$or": [
            {"data.fabric_types": capability_pattern}, {"data.compositions": capability_pattern},
            {"data.shade_colours": capability_pattern}, {"data.services": capability_pattern},
            {"data.quality_standards": capability_pattern}, {"data.bulk_price_note": capability_pattern},
            {"data.stock_availability": capability_pattern}, {"data.brands": capability_pattern},
            {"data.sales_channels": capability_pattern}, {"data.export_countries": capability_pattern},
            {"data.incoterms": capability_pattern}, {"data.currencies": capability_pattern},
            {"data.export_documents": capability_pattern}, {"data.store_categories": capability_pattern},
            {"data.seasonal_requirements": capability_pattern},
        ]})
    if service_region and service_region.strip():
        region_pattern = {"$regex": service_region.strip(), "$options": "i"}
        profile_filters.append({"$or": [
            {"data.service_regions": region_pattern}, {"data.territories": region_pattern},
            {"data.sourcing_regions": region_pattern}, {"data.delivery_locations": region_pattern},
        ]})
    if profile_filters:
        profile_query["$and"] = profile_filters

    vendor_operations: dict = {}
    async for operation in vendor_role_operations_collection.find(profile_query, {"vendor_id": 1, "role": 1, "data": 1}):
        vendor_operations.setdefault(operation["vendor_id"], []).append({
            "role": operation.get("role", ""),
            "data": operation.get("data", {}),
        })

    if profile_filters:
        vendor_map = {
            vendor_id: details for vendor_id, details in vendor_map.items()
            if vendor_id in vendor_operations
        }
        if not vendor_map:
            return {"status": "success", "count": 0, "data": []}

    item_query: dict = {"vendor_id": {"$in": list(vendor_map.keys())}, "active": True}
    text_filters = []
    for term in (category, material, audience):
        if term and term.strip():
            text_filters.append({"$or": [
                {"item_name":   {"$regex": term.strip(), "$options": "i"}},
                {"category":    {"$regex": term.strip(), "$options": "i"}},
                {"description": {"$regex": term.strip(), "$options": "i"}},
            ]})
    if text_filters:
        item_query["$and"] = text_filters
    if min_price is not None:
        item_query["price_range_max"] = {"$gte": min_price}
    if max_price is not None:
        item_query.setdefault("price_range_min", {})["$lte"] = max_price

    results = []
    async for item in vendor_catalogue_collection.find(item_query).sort("created_at", -1):
        row = _serialize_item(item)
        vinfo = vendor_map.get(item["vendor_id"], {})
        row["vendor_name"]    = vinfo.get("vendor_name", "")
        row["business_type"]  = vinfo.get("business_type", [])
        row["supplier_operations"] = vendor_operations.get(item["vendor_id"], [])
        row["featured_badge"] = vinfo.get("featured_badge", False)
        row["_priority_rank"] = vinfo.get("priority_rank", 0)
        results.append(row)

    # Premium (rank 2) → Standard (rank 1) → Free (rank 0), newest first
    # within each tier — this IS the paid placement benefit, not a
    # cosmetic sort.
    results.sort(key=lambda r: (-r["_priority_rank"], ))
    for r in results:
        del r["_priority_rank"]

    return {"status": "success", "count": len(results), "data": results}


@router.post("/my-catalogue/{item_id}/share")
async def share_catalogue_item_with_retailers(item_id: str, payload: dict, authorization: str = Header(None)):
    """Notify selected approved retailer relationships about one active catalogue item."""
    vendor_id = _decode_vendor(authorization)
    if not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="Invalid catalogue item ID.")
    item = await vendor_catalogue_collection.find_one({"_id": ObjectId(item_id), "vendor_id": ObjectId(vendor_id), "active": True})
    if not item:
        raise HTTPException(status_code=404, detail="Only an active item in your catalogue can be shared.")
    requested_tenants = list(dict.fromkeys(str(value).strip() for value in (payload.get("tenant_ids") or []) if str(value).strip()))
    if not requested_tenants:
        raise HTTPException(status_code=400, detail="Select at least one approved retailer.")
    approved_links = await vendor_tenant_links_collection.find({
        "vendor_id": ObjectId(vendor_id), "tenant_id": {"$in": requested_tenants}, "status": "Approved",
    }).to_list(length=None)
    approved_tenants = {link.get("tenant_id") for link in approved_links}
    if not approved_tenants:
        raise HTTPException(status_code=403, detail="You can share only with retailers that have approved your vendor account.")
    vendor = await vendors_collection.find_one({"_id": ObjectId(vendor_id)}, {"name": 1, "vendor_name": 1, "brandName": 1})
    vendor_name = (vendor or {}).get("brandName") or (vendor or {}).get("name") or (vendor or {}).get("vendor_name") or "A vendor"
    item_name = item.get("item_name") or "a catalogue item"
    for tenant_id in approved_tenants:
        await notify_buyer(
            tenant_id, "catalogue_item_shared", "New vendor catalogue item shared",
            f"{vendor_name} shared {item_name}. Open Quick Order to select variants and request a quote.",
            vendor_id=ObjectId(vendor_id),
            metadata={"catalogue_item_id": str(item["_id"]), "vendor_id": str(vendor_id), "item_name": item_name, "shared_by": vendor_name},
        )
    return {"status": "success", "message": f"Shared with {len(approved_tenants)} retailer(s).", "shared_tenant_ids": sorted(approved_tenants)}

# ═══════════════════════════════════════════════════════════════════════════
# INQUIRIES — the buyer-vendor negotiation loop
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/shared-item/{item_id}")
async def get_shared_catalogue_item(item_id: str, ctx: dict = Depends(get_hq_tenant)):
    """Return one shareable listing only when this retailer has approved the vendor."""
    if not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="Invalid catalogue item ID.")
    item = await vendor_catalogue_collection.find_one({"_id": ObjectId(item_id), "active": True})
    if not item:
        raise HTTPException(status_code=404, detail="Shared catalogue item is unavailable.")
    link = await vendor_tenant_links_collection.find_one({"vendor_id": item["vendor_id"], "tenant_id": ctx["tenant_id"], "status": "Approved"})
    if not link:
        raise HTTPException(status_code=403, detail="This retailer is not approved to view this vendor item.")
    return {"status": "success", "data": _serialize_item(item)}

@router.post("/inquiries", status_code=201)
async def create_inquiry(payload: dict, ctx: dict = Depends(get_hq_tenant)):
    """
    Buyer asks about a specific catalogue item — price/size/color/qty they
    want. Vendor sees this and responds; nothing is committed to a PO yet.
    """
    item_id = payload.get("catalogue_item_id")
    if not item_id or not ObjectId.is_valid(item_id):
        raise HTTPException(status_code=400, detail="Valid catalogue_item_id is required.")

    item = await vendor_catalogue_collection.find_one({"_id": ObjectId(item_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Catalogue item not found.")

    # Re-verify the approved relationship — a buyer shouldn't be able to
    # inquire about a vendor's catalogue they don't actually have access to,
    # even if they somehow have the item_id.
    approved_link = await vendor_tenant_links_collection.find_one({
        "vendor_id": item["vendor_id"],
        "tenant_id": ctx["tenant_id"],
        "status":    "Approved",
    })
    if not approved_link:
        raise HTTPException(status_code=403, detail="You don't have an approved relationship with this vendor.")

    doc = {
        "catalogue_item_id": item["_id"],
        "vendor_id":         item["vendor_id"],
        "tenant_id":         ctx["tenant_id"],
        "item_name":         item.get("item_name", ""),
        "item_image":        (item.get("images") or [None])[0],
        "requested_size":    (payload.get("requested_size") or "").strip(),
        "requested_color":   (payload.get("requested_color") or "").strip(),
        "requested_qty":     max(0, int(payload.get("requested_qty") or 0)),
        "requested_price":   max(0.0, float(payload.get("requested_price") or 0)),
        "buyer_note":        (payload.get("buyer_note") or "").strip(),
        "response_deadline": (payload.get("response_deadline") or "").strip(),
        "status":            "Pending",   # Pending | Responded | Converted | Declined
        "created_by":        ctx.get("admin_id"),
        "created_at":        datetime.utcnow(),
        "vendor_response":   None,
    }
    result = await catalogue_inquiries_collection.insert_one(doc)
    await notify_vendor(item["vendor_id"], "rfq_received", "New buyer inquiry", f"A buyer requested a quote for {item.get('item_name', 'a catalogue item')}.", tenant_id=ctx["tenant_id"], inquiry_id=result.inserted_id)
    return {"status": "success", "message": "Inquiry sent to vendor.", "id": str(result.inserted_id)}


@router.get("/inquiries")
async def list_my_inquiries(ctx: dict = Depends(get_hq_tenant)):
    """
    Buyer's own tenant's inquiries, across all vendors.

    ⚠️ FIXED: previously returned rows with NO vendor_name at all — the
    raw catalogue_inquiries_collection document only ever stored
    vendor_id, never a name, and this route never joined one in. A buyer
    browsing "My Inquiries" had no way to tell which vendor a given
    Responded quote actually came from. Same fix pattern already applied
    to the vendor-side list_vendor_inquiries (which joins tenant_name) —
    this is the mirror fix for the buyer side.
    """
    docs = await catalogue_inquiries_collection.find(
        {"tenant_id": ctx["tenant_id"]}
    ).sort("created_at", -1).to_list(None)

    vendor_ids = list({d["vendor_id"] for d in docs})
    vendor_map: dict = {}
    if vendor_ids:
        async for v in vendors_collection.find({"_id": {"$in": vendor_ids}}):
            vendor_map[v["_id"]] = v.get("name") or v.get("vendor_name", "Unknown Vendor")

    rows = []
    for doc in docs:
        row = _serialize_inquiry(doc)
        row["vendor_name"] = vendor_map.get(doc["vendor_id"], "Unknown Vendor")
        rows.append(row)
    return {"status": "success", "data": rows}


@router.get("/my-inquiries")
async def list_vendor_inquiries(authorization: str = Header(None)):
    """
    Vendor's own inquiries, across every retailer that's asked — each row
    carries tenant_id + a resolved company_name so the vendor knows who's
    asking.

    ⚠️ FIXED: previously only showed the buyer's requested price/size/
    color — the vendor had no way to see their OWN listed price range or
    available colors on this same screen while deciding how to respond,
    without switching to the Catalogue tab separately. Now joins each
    inquiry's original catalogue_item_id back to vendor_catalogue_collection
    and attaches my_price_range_min/max and my_available_colors/sizes —
    the vendor's own reference point to negotiate against.
    """
    vendor_id = _decode_vendor(authorization)

    name_map: dict = {}
    async for t in tenants_collection.find({}, {"_id": 0, "tenant_id": 1, "company_name": 1}):
        if t.get("tenant_id"):
            name_map[t["tenant_id"]] = t.get("company_name") or t["tenant_id"]

    # Pre-fetch the vendor's own catalogue items once, keyed by _id, rather
    # than a separate DB call per inquiry row.
    catalogue_map: dict = {}
    async for item in vendor_catalogue_collection.find({"vendor_id": ObjectId(vendor_id)}):
        catalogue_map[item["_id"]] = item

    rows = []
    async for doc in catalogue_inquiries_collection.find(
        {"vendor_id": ObjectId(vendor_id)}
    ).sort("created_at", -1):
        row = _serialize_inquiry(doc)
        row["tenant_name"] = name_map.get(doc.get("tenant_id"), doc.get("tenant_id"))

        source_item = catalogue_map.get(doc.get("catalogue_item_id"))
        if source_item:
            row["my_price_range_min"]  = source_item.get("price_range_min", 0)
            row["my_price_range_max"]  = source_item.get("price_range_max", 0)
            row["my_available_colors"] = source_item.get("available_colors", [])
            row["my_available_sizes"]  = source_item.get("available_sizes", [])
        else:
            # Catalogue item was deleted since this inquiry was raised —
            # don't crash, just show nothing rather than a stale/wrong price.
            row["my_price_range_min"]  = None
            row["my_price_range_max"]  = None
            row["my_available_colors"] = []
            row["my_available_sizes"]  = []

        rows.append(row)
    return {"status": "success", "data": rows}


@router.get("/analytics")
async def get_vendor_analytics(authorization: str = Header(None)):
    """
    Inquiry performance stats — total received, response rate, conversion
    rate, and a breakdown by retailer. Built entirely from data that
    already exists in catalogue_inquiries_collection; no new tracking or
    instrumentation needed.

    Gated to Standard/Premium — Free vendors get a paywall message instead
    of the numbers. This is a genuine reason to pay beyond "more storage":
    a vendor with real inquiry volume can see whether they're actually
    converting, which Free vendors currently have no visibility into at
    all (they can only see individual inquiry statuses one at a time on
    the Inquiries tab, not aggregated).
    """
    vendor_id = _decode_vendor(authorization)
    tier = await get_vendor_tier(vendor_id)

    if tier["tier"] == "free":
        raise HTTPException(
            status_code=403,
            detail="Analytics are available on Standard and Premium plans. Upgrade to see your inquiry performance."
        )

    all_inquiries = []
    async for doc in catalogue_inquiries_collection.find({"vendor_id": ObjectId(vendor_id)}):
        all_inquiries.append(doc)

    total = len(all_inquiries)
    responded = sum(1 for i in all_inquiries if i["status"] in ("Responded", "Converted"))
    converted = sum(1 for i in all_inquiries if i["status"] == "Converted")
    declined  = sum(1 for i in all_inquiries if i["status"] == "Declined")

    by_tenant: dict = {}
    for i in all_inquiries:
        t = i.get("tenant_id", "unknown")
        by_tenant[t] = by_tenant.get(t, 0) + 1

    return {
        "status": "success",
        "data": {
            "total_inquiries":     total,
            "responded":           responded,
            "converted":           converted,
            "declined":            declined,
            "response_rate_pct":   round((responded / total) * 100, 1) if total else 0,
            "conversion_rate_pct": round((converted / total) * 100, 1) if total else 0,
            "inquiries_by_retailer_count": len(by_tenant),
        },
    }


@router.post("/inquiries/{inquiry_id}/respond")
async def respond_to_inquiry(inquiry_id: str, payload: dict, authorization: str = Header(None)):
    """Vendor responds to a buyer's inquiry with a firm offer."""
    vendor_id = _decode_vendor(authorization)
    if not ObjectId.is_valid(inquiry_id):
        raise HTTPException(status_code=400, detail="Invalid inquiry ID.")

    inquiry = await catalogue_inquiries_collection.find_one({
        "_id": ObjectId(inquiry_id), "vendor_id": ObjectId(vendor_id),
    })
    if not inquiry:
        raise HTTPException(status_code=404, detail="Inquiry not found.")
    if inquiry.get("status") in ("Converted", "Closed", "Cancelled", "Expired"):
        raise HTTPException(status_code=400, detail=f"This inquiry is {inquiry.get('status')} and cannot receive quotations.")
    deadline = inquiry.get("response_deadline") or ""
    if deadline and deadline < datetime.utcnow().strftime("%Y-%m-%d"):
        await catalogue_inquiries_collection.update_one({"_id": inquiry["_id"]}, {"$set": {"status": "Expired", "expired_at": datetime.utcnow()}})
        raise HTTPException(status_code=400, detail="The RFQ response deadline has expired.")

    available       = bool(payload.get("available", True))
    confirmed_price = max(0.0, float(payload.get("confirmed_price") or 0))
    confirmed_qty   = max(0, int(payload.get("confirmed_qty") or inquiry.get("requested_qty") or 0))

    # ⚠️ Fix: previously a vendor could mark "available" with no price at
    # all, defaulting silently to ₹0 — that ₹0 would then be convertible
    # straight into a real PO line item. If the vendor is confirming
    # availability, a real price and quantity are required; otherwise they
    # must explicitly decline instead.
    if available and (confirmed_price <= 0 or confirmed_qty <= 0):
        raise HTTPException(
            status_code=400,
            detail="A confirmed price and quantity greater than zero are required to mark this available."
        )

    vendor_response = {
        "confirmed_size":  (payload.get("confirmed_size") or inquiry.get("requested_size", "")).strip(),
        "confirmed_color": (payload.get("confirmed_color") or inquiry.get("requested_color", "")).strip(),
        "confirmed_qty":   confirmed_qty,
        "confirmed_price": confirmed_price,
        "available":       available,
        "vendor_note":     (payload.get("vendor_note") or "").strip(),
        "discount_pct":    min(100.0, max(0.0, float(payload.get("discount_pct") or 0))),
        "tax_pct":         min(100.0, max(0.0, float(payload.get("tax_pct") or 0))),
        "freight":         max(0.0, float(payload.get("freight") or 0)),
        "other_charges":   max(0.0, float(payload.get("other_charges") or 0)),
        "payment_terms":   (payload.get("payment_terms") or "").strip(),
        "credit_days":     max(0, int(payload.get("credit_days") or 0)),
        "lead_time_days":  max(0, int(payload.get("lead_time_days") or 0)),
        "delivery_date":   (payload.get("delivery_date") or "").strip(),
        "sample_cost":     max(0.0, float(payload.get("sample_cost") or 0)),
        "moq":             max(0, int(payload.get("moq") or 0)),
        "quote_valid_until": (payload.get("quote_valid_until") or "").strip(),
    }
    base_value = confirmed_qty * confirmed_price
    discount_amount = base_value * vendor_response["discount_pct"] / 100
    taxable_value = base_value - discount_amount
    tax_amount = taxable_value * vendor_response["tax_pct"] / 100
    vendor_response.update({
        "base_value": base_value, "discount_amount": discount_amount,
        "tax_amount": tax_amount,
        "landed_cost": taxable_value + tax_amount + vendor_response["freight"] + vendor_response["other_charges"],
    })

    now = datetime.utcnow()
    revision_number = len(inquiry.get("quotation_revisions") or []) + 1
    quotation_revision = {
        **vendor_response,
        "revision": revision_number,
        "created_at": now,
        "created_by": "vendor",
    }
    event = {
        "type": "vendor_quote" if available else "vendor_declined",
        "actor": "vendor",
        "message": vendor_response.get("vendor_note") or (f"Quotation revision {revision_number}" if available else "Vendor declined"),
        "price": confirmed_price,
        "quantity": confirmed_qty,
        "created_at": now,
    }
    await catalogue_inquiries_collection.update_one(
        {"_id": ObjectId(inquiry_id)},
        {
            "$set": {
                "status":          "Responded" if available else "Declined",
                "vendor_response": vendor_response,
                "responded_at":    now,
            },
            "$push": {
                "quotation_revisions": quotation_revision,
                "negotiation_history": event,
            },
        },
    )
    await notify_buyer(inquiry["tenant_id"], "quote_revised", "Vendor quotation received", f"A vendor submitted quotation revision {revision_number} for {inquiry.get('item_name', 'your inquiry')}.", inquiry_id=inquiry["_id"], vendor_id=inquiry["vendor_id"], metadata={"revision": revision_number, "price": confirmed_price})
    return {"status": "success", "message": "Response sent to buyer.", "revision": revision_number}


async def _change_inquiry_lifecycle(inquiry_id: str, tenant_id: str, target: str, allowed: tuple, admin_id=None):
    if not ObjectId.is_valid(inquiry_id): raise HTTPException(status_code=400, detail="Invalid inquiry ID.")
    inquiry = await catalogue_inquiries_collection.find_one({"_id": ObjectId(inquiry_id), "tenant_id": tenant_id})
    if not inquiry: raise HTTPException(status_code=404, detail="Inquiry not found.")
    if inquiry.get("status") not in allowed: raise HTTPException(status_code=409, detail=f"Cannot change {inquiry.get('status')} inquiry to {target}.")
    field = {"Closed":"closed_at", "Cancelled":"cancelled_at", "Pending":"reopened_at"}[target]
    await catalogue_inquiries_collection.update_one({"_id": inquiry["_id"], "tenant_id": tenant_id}, {"$set": {"status": target, field: datetime.utcnow(), "lifecycle_changed_by": admin_id}})
    return {"status": "success", "inquiry_status": target}


@router.post("/inquiries/{inquiry_id}/close")
async def close_inquiry(inquiry_id: str, ctx: dict = Depends(get_hq_tenant)):
    return await _change_inquiry_lifecycle(inquiry_id, ctx["tenant_id"], "Closed", ("Pending","Responded","Countered","Expired"), ctx.get("admin_id"))


@router.post("/inquiries/{inquiry_id}/cancel")
async def cancel_inquiry(inquiry_id: str, ctx: dict = Depends(get_hq_tenant)):
    return await _change_inquiry_lifecycle(inquiry_id, ctx["tenant_id"], "Cancelled", ("Pending","Responded","Countered","Expired","Closed"), ctx.get("admin_id"))


@router.post("/inquiries/{inquiry_id}/reopen")
async def reopen_inquiry(inquiry_id: str, payload: dict, ctx: dict = Depends(get_hq_tenant)):
    result = await _change_inquiry_lifecycle(inquiry_id, ctx["tenant_id"], "Pending", ("Closed","Cancelled","Expired"), ctx.get("admin_id"))
    deadline=(payload.get("response_deadline") or "").strip()
    if deadline: await catalogue_inquiries_collection.update_one({"_id":ObjectId(inquiry_id),"tenant_id":ctx["tenant_id"]},{"$set":{"response_deadline":deadline}})
    return result


@router.post("/inquiries/{inquiry_id}/counteroffer")
async def create_buyer_counteroffer(inquiry_id: str, payload: dict, ctx: dict = Depends(get_hq_tenant)):
    """Tenant-scoped buyer counteroffer; preserves the vendor quote and requests a revision."""
    if not ObjectId.is_valid(inquiry_id):
        raise HTTPException(status_code=400, detail="Invalid inquiry ID.")
    inquiry = await catalogue_inquiries_collection.find_one({
        "_id": ObjectId(inquiry_id), "tenant_id": ctx["tenant_id"],
    })
    if not inquiry:
        raise HTTPException(status_code=404, detail="Inquiry not found.")
    if inquiry.get("status") not in ("Responded", "Countered"):
        raise HTTPException(status_code=400, detail="A counteroffer can only be sent after a vendor quotation.")
    valid_until = (inquiry.get("vendor_response") or {}).get("quote_valid_until") or ""
    if valid_until and valid_until < datetime.utcnow().strftime("%Y-%m-%d"):
        raise HTTPException(status_code=400, detail="The vendor quotation has expired; request a revised quotation.")
    message = (payload.get("message") or "").strip()
    target_price = max(0.0, float(payload.get("target_price") or 0))
    target_qty = max(0, int(payload.get("target_qty") or inquiry.get("requested_qty") or 0))
    if not message and target_price <= 0:
        raise HTTPException(status_code=400, detail="Enter a target price or negotiation message.")
    now = datetime.utcnow()
    event = {
        "type": "buyer_counteroffer", "actor": "buyer", "message": message,
        "price": target_price, "quantity": target_qty, "created_at": now,
        "created_by": ctx.get("admin_id"),
    }
    await catalogue_inquiries_collection.update_one(
        {"_id": ObjectId(inquiry_id), "tenant_id": ctx["tenant_id"]},
        {"$set": {"status": "Countered", "countered_at": now}, "$push": {"negotiation_history": event}},
    )
    await notify_vendor(inquiry["vendor_id"], "counteroffer_received", "Buyer counteroffer received", f"The buyer sent a counteroffer for {inquiry.get('item_name', 'your quotation')}.", tenant_id=ctx["tenant_id"], inquiry_id=inquiry["_id"], metadata={"target_price": target_price, "target_qty": target_qty})
    return {"status": "success", "message": "Counteroffer sent to vendor."}


def _build_export_rows(inquiries: list, vendor_map: dict) -> list:
    rows = []
    for inq in inquiries:
        vr = inq.get("vendor_response") or {}
        qty   = vr.get("confirmed_qty", 0)
        price = vr.get("confirmed_price", 0)
        rows.append({
            "vendor":   vendor_map.get(inq["vendor_id"], "Unknown Vendor"),
            "item":     inq.get("item_name", ""),
            "size":     vr.get("confirmed_size", "") or "—",
            "color":    vr.get("confirmed_color", "") or "—",
            "qty":      qty,
            "price":    price,
            "subtotal": qty * price,
        })
    return rows


def _export_xlsx(rows: list, tenant_id: str) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation Summary"

    headers = ["Vendor", "Item", "Size", "Color", "Qty", "Price (₹)", "Subtotal (₹)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    grand_total = 0.0
    for r in rows:
        ws.append([r["vendor"], r["item"], r["size"], r["color"], r["qty"], r["price"], r["subtotal"]])
        grand_total += r["subtotal"]

    ws.append([])
    total_row = ["", "", "", "", "", "Grand Total", round(grand_total, 2)]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"quotation_summary_{tenant_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_pdf(rows: list, tenant_id: str) -> StreamingResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleX", parent=styles["Heading1"], textColor=colors.HexColor("#4F46E5"))

    elements = [
        Paragraph("Vendor Quotation Summary", title_style),
        Paragraph(f"Retailer: {tenant_id} &nbsp;·&nbsp; Generated: {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}", styles["Normal"]),
        Spacer(1, 10 * mm),
    ]

    data = [["Vendor", "Item", "Size", "Color", "Qty", "Price (₹)", "Subtotal (₹)"]]
    grand_total = 0.0
    for r in rows:
        data.append([r["vendor"], r["item"], r["size"], r["color"], str(r["qty"]), f"{r['price']:.2f}", f"{r['subtotal']:.2f}"])
        grand_total += r["subtotal"]
    data.append(["", "", "", "", "", "Grand Total", f"{grand_total:.2f}"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
        ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME",       (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN",          (4, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F3F4F6")]),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",     (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    filename = f"quotation_summary_{tenant_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/inquiries/export")
async def export_inquiries(payload: dict, ctx: dict = Depends(get_hq_tenant)):
    """
    Buyer selects a set of negotiated inquiries (Responded or already
    Converted) and downloads a summary sheet — PDF or Excel — of item,
    size, color, qty, negotiated price, and subtotals across possibly
    multiple vendors.

    Read-only: does NOT change any inquiry's status, so it can be
    regenerated any number of times (e.g. re-download after already
    converting some items to a PO, for record-keeping).
    """
    inquiry_ids = payload.get("inquiry_ids", [])
    award_quantities = payload.get("award_quantities") or {}
    fmt = (payload.get("format") or "pdf").lower()

    if not inquiry_ids:
        raise HTTPException(status_code=400, detail="At least one inquiry_id is required.")
    if fmt not in ("pdf", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be 'pdf' or 'xlsx'.")

    oids = []
    for iid in inquiry_ids:
        if not ObjectId.is_valid(iid):
            raise HTTPException(status_code=400, detail=f"Invalid inquiry ID: {iid}")
        oids.append(ObjectId(iid))

    inquiries = await catalogue_inquiries_collection.find({
        "_id": {"$in": oids},
        "tenant_id": ctx["tenant_id"],
        "status": {"$in": ["Responded", "Converted"]},
    }).to_list(None)

    if not inquiries:
        raise HTTPException(
            status_code=404,
            detail="No valid Responded/Converted inquiries found for the given IDs. "
                   "Only negotiated (vendor-confirmed) inquiries can be exported."
        )

    vendor_ids = list({i["vendor_id"] for i in inquiries})
    vendor_map: dict = {}
    async for v in vendors_collection.find({"_id": {"$in": vendor_ids}}):
        vendor_map[v["_id"]] = v.get("name") or v.get("vendor_name") or "Unknown Vendor"

    # Comparison exports may carry buyer-awarded split quantities. Never
    # allow an export quantity above the vendor-confirmed availability.
    export_inquiries = []
    for inquiry in inquiries:
        row = dict(inquiry)
        response = dict(row.get("vendor_response") or {})
        requested_award = award_quantities.get(str(row["_id"]))
        if requested_award is not None:
            confirmed = max(0, int(response.get("confirmed_qty") or 0))
            response["confirmed_qty"] = min(max(0, int(requested_award or 0)), confirmed)
            row["vendor_response"] = response
        export_inquiries.append(row)
    rows = _build_export_rows(export_inquiries, vendor_map)

    if fmt == "xlsx":
        return _export_xlsx(rows, ctx["tenant_id"])
    return _export_pdf(rows, ctx["tenant_id"])


@router.post("/inquiries/{inquiry_id}/convert")
async def convert_inquiry(inquiry_id: str, ctx: dict = Depends(get_hq_tenant)):
    """
    Buyer converts a Responded inquiry into prefill data for the PO
    creation form. Does NOT create the PO itself — returns the resolved
    line-item data (description, size, color, qty, rate) for the frontend
    to drop straight into an item row, and marks the inquiry Converted so
    it can't be reused.
    """
    if not ObjectId.is_valid(inquiry_id):
        raise HTTPException(status_code=400, detail="Invalid inquiry ID.")

    inquiry = await catalogue_inquiries_collection.find_one({
        "_id": ObjectId(inquiry_id), "tenant_id": ctx["tenant_id"],
    })
    if not inquiry:
        raise HTTPException(status_code=404, detail="Inquiry not found.")
    if inquiry.get("status") != "Responded":
        raise HTTPException(status_code=400, detail="Only a Responded inquiry can be converted to a PO.")

    vr = inquiry.get("vendor_response") or {}

    # Defense in depth: even though respond_to_inquiry already requires a
    # real price/qty before allowing "Responded", re-check here too rather
    # than trusting that state can only ever be reached one way.
    if not vr.get("confirmed_price") or vr["confirmed_price"] <= 0 or not vr.get("confirmed_qty") or vr["confirmed_qty"] <= 0:
        raise HTTPException(
            status_code=400,
            detail="This inquiry has no valid confirmed price/quantity and cannot be converted."
        )

    await catalogue_inquiries_collection.update_one(
        {"_id": ObjectId(inquiry_id)},
        {"$set": {"status": "Converted", "converted_at": datetime.utcnow()}}
    )

    vendor = await vendors_collection.find_one({"_id": inquiry["vendor_id"]})

    return {
        "status": "success",
        "po_item_prefill": {
            "description": f"{inquiry.get('item_name','')} — {vr.get('confirmed_size','')} / {vr.get('confirmed_color','')}".strip(" —/"),
            "quantity":    vr.get("confirmed_qty", 0),
            "rate":        vr.get("confirmed_price", 0),
            "remarks":     vr.get("vendor_note", ""),
        },
        "vendor_name": (vendor or {}).get("name") or (vendor or {}).get("vendor_name", ""),
        "vendor_id":   str(inquiry["vendor_id"]),
    }


# ═══════════════════════════════════════════════════════════════════════════
# CATALOGUE EXPIRY SWEEP — the mechanism behind "images visible for 45/90 days"
# ═══════════════════════════════════════════════════════════════════════════

async def purge_expired_catalogue_media(now: Optional[datetime] = None) -> dict:
    """Permanently remove only Cloudinary media for listings hidden 30+ days.

    Catalogue items, variants, inquiry history, orders and invoices remain.
    A vendor can add fresh images and reactivate the same item later. If a
    Cloudinary deletion fails, the item is retained for a safe retry.
    """
    now = now or datetime.utcnow()
    cutoff = now - timedelta(days=CATALOGUE_MEDIA_RETENTION_DAYS)
    candidates = await vendor_catalogue_collection.find({
        "active": False,
        "expired_reason": "tier_visibility_window",
        "expired_at": {"$lte": cutoff},
        "media_purged_at": {"$exists": False},
    }).to_list(length=None)
    purged_items = 0
    purged_images = 0
    failed_items = 0
    for item in candidates:
        image_urls = list(item.get("images") or [])
        delete_failed = False
        for image_url in image_urls:
            public_id = _vendor_catalogue_public_id(str(image_url))
            if not public_id:
                continue
            try:
                result = cloudinary.uploader.destroy(public_id, resource_type="image", invalidate=True)
                if result.get("result") not in {"ok", "not found"}:
                    delete_failed = True
                    break
            except Exception as exc:
                print(f"Catalogue media cleanup failed for {public_id}: {exc}")
                delete_failed = True
                break
        if delete_failed:
            failed_items += 1
            continue
        result = await vendor_catalogue_collection.update_one(
            {
                "_id": item["_id"], "active": False,
                "expired_reason": "tier_visibility_window",
                "expired_at": {"$lte": cutoff},
                "media_purged_at": {"$exists": False},
            },
            {"$set": {
                "images": [], "media_purged_at": now,
                "media_purge_reason": "retention_period_elapsed",
                "purged_image_count": len(image_urls), "updated_at": now,
            }},
        )
        if result.modified_count:
            purged_items += 1
            purged_images += len(image_urls)
    return {"media_purged_items": purged_items, "media_purged_images": purged_images, "media_cleanup_failed_items": failed_items}


async def expire_stale_catalogue_items() -> dict:
    """
    Soft-hides (active=False) every catalogue item whose expires_at has
    passed. Does NOT delete anything — a vendor renewing or upgrading later
    should be able to see their history, and a hard delete would also
    orphan any inquiry that still references the item.

    This is a plain async function, not just a route, so it can be called
    two ways:
      1. Via the POST /api/catalogue/expire-sweep route below (for manual
         triggering / testing).
      2. Directly from a real scheduler in your own code, e.g.:
             from app.routes.catalogue_routes import expire_stale_catalogue_items
             scheduler.add_job(expire_stale_catalogue_items, "cron", hour=2)
         using APScheduler, or a Celery beat task, or literally anything
         that can call an async Python function once a day. NOTHING calls
         this automatically right now — without a real scheduler wired up,
         catalogue items will never actually expire no matter what
         expires_at says.
    """
    now = datetime.utcnow()
    result = await vendor_catalogue_collection.update_many(
        {"active": True, "expires_at": {"$lt": now}},
        {"$set": {"active": False, "expired_at": now, "expired_reason": "tier_visibility_window"}}
    )
    media_cleanup = await purge_expired_catalogue_media(now)
    return {"expired_count": result.modified_count, "swept_at": now.isoformat(), **media_cleanup}


@router.post("/expire-sweep")
async def run_expire_sweep(x_cron_secret: Optional[str] = Header(None)):
    """
    Manually triggers expire_stale_catalogue_items(). Protected by a shared
    secret header rather than vendor/HQ auth, since this is meant to be
    called by a scheduler/cron job, not a logged-in user.

    Set CRON_SECRET in your environment and configure whatever scheduler
    you use to send it as the X-Cron-Secret header. Without CRON_SECRET
    set, this route refuses all requests — it does NOT fail open.
    """
    expected = os.environ.get("CRON_SECRET")
    if not expected:
        raise HTTPException(
            status_code=503,
            detail="CRON_SECRET is not configured on the server — this route is disabled until it is."
        )
    if x_cron_secret != expected:
        raise HTTPException(status_code=401, detail="Invalid or missing X-Cron-Secret header.")

    result = await expire_stale_catalogue_items()
    return {"status": "success", **result}


# ═══════════════════════════════════════════════════════════════════════════
'''
@router.get("/approved-vendors")
async def list_approved_vendors_for_rfq(ctx: dict = Depends(get_hq_tenant)):
    """Approved vendors available to this retailer for an open RFQ."""
    vendor_ids = []
    async for link in vendor_tenant_links_collection.find({
        "tenant_id": ctx["tenant_id"], "status": "Approved",
    }, {"vendor_id": 1}):
        vendor_ids.append(link["vendor_id"])
    if not vendor_ids:
        return {"status": "success", "data": []}
    rows = []
    async for vendor in vendors_collection.find({"_id": {"$in": vendor_ids}}):
        rows.append({
            "_id": str(vendor["_id"]),
            "vendor_name": vendor.get("name") or vendor.get("vendor_name") or "Unnamed vendor",
            "business_type": vendor.get("business_type") or [],
            "city": vendor.get("city") or vendor.get("address", {}).get("city", ""),
        })
    rows.sort(key=lambda row: row["vendor_name"].lower())
    return {"status": "success", "data": rows}
'''


@router.get("/approved-vendors")
async def list_approved_vendors_for_rfq(ctx: dict = Depends(get_hq_tenant)):
    """Approved vendors available to this retailer for an open RFQ."""
    vendor_ids = []
    async for link in vendor_tenant_links_collection.find({
        "tenant_id": ctx["tenant_id"], "status": "Approved",
    }, {"vendor_id": 1}):
        vendor_ids.append(link["vendor_id"])
    if not vendor_ids:
        return {"status": "success", "data": []}
    rows = []
    async for vendor in vendors_collection.find({"_id": {"$in": vendor_ids}}):
        address = vendor.get("address")
        # ⚠️ FIX: address can be stored either as a nested object
        # ({"city": "...", ...}) or as a plain free-text string
        # ("123 MG Road, Pune") depending on how/when the vendor record
        # was created. Only call .get() on it when it's actually a dict.
        address_city = address.get("city", "") if isinstance(address, dict) else ""
        rows.append({
            "_id": str(vendor["_id"]),
            "vendor_name": vendor.get("name") or vendor.get("vendor_name") or "Unnamed vendor",
            "business_type": vendor.get("business_type") or [],
            "city": vendor.get("city") or address_city,
        })
    rows.sort(key=lambda row: row["vendor_name"].lower())
    return {"status": "success", "data": rows}

@router.post("/inquiries/open-rfq", status_code=201)
async def create_open_rfq(payload: dict, ctx: dict = Depends(get_hq_tenant)):
    """Send one non-catalogue sourcing requirement to approved vendors."""
    vendor_ids = list(dict.fromkeys(payload.get("vendor_ids") or []))
    item_name = (payload.get("item_name") or "").strip()
    requested_qty = max(0, int(payload.get("requested_qty") or 0))
    if not item_name:
        raise HTTPException(status_code=400, detail="Product or requirement name is required.")
    if requested_qty <= 0:
        raise HTTPException(status_code=400, detail="Requested quantity must be greater than zero.")
    if not vendor_ids:
        raise HTTPException(status_code=400, detail="Select at least one vendor.")

    valid_oids = [ObjectId(v) for v in vendor_ids if ObjectId.is_valid(v)]
    approved = set()
    async for link in vendor_tenant_links_collection.find({
        "tenant_id": ctx["tenant_id"], "status": "Approved", "vendor_id": {"$in": valid_oids},
    }, {"vendor_id": 1}):
        approved.add(link["vendor_id"])
    if not approved:
        raise HTTPException(status_code=403, detail="None of the selected vendors are approved for this retailer.")

    group_id = str(uuid.uuid4())
    created = []
    now = datetime.utcnow()
    for vendor_id in approved:
        doc = {
            "catalogue_item_id": None,
            "vendor_id": vendor_id,
            "tenant_id": ctx["tenant_id"],
            "inquiry_type": "open_rfq",
            "item_name": item_name,
            "item_image": payload.get("reference_image_url") or None,
            "category": (payload.get("category") or "").strip(),
            "material": (payload.get("material") or "").strip(),
            "audience": (payload.get("audience") or "").strip(),
            "requested_size": (payload.get("requested_size") or "").strip(),
            "requested_color": (payload.get("requested_color") or "").strip(),
            "requested_qty": requested_qty,
            "requested_price": max(0.0, float(payload.get("target_price") or 0)),
            "price_range_min": max(0.0, float(payload.get("price_range_min") or 0)),
            "price_range_max": max(0.0, float(payload.get("price_range_max") or 0)),
            "delivery_date": (payload.get("delivery_date") or "").strip(),
            "buyer_note": (payload.get("buyer_note") or "").strip(),
            "response_deadline": (payload.get("response_deadline") or "").strip(),
            "allow_alternatives": bool(payload.get("allow_alternatives", True)),
            "status": "Pending",
            "comparison_group_id": group_id,
            "created_by": ctx.get("admin_id"),
            "created_at": now,
            "vendor_response": None,
        }
        result = await catalogue_inquiries_collection.insert_one(doc)
        await notify_vendor(vendor_id, "open_rfq_received", "New open RFQ", f"A buyer is sourcing {item_name}.", tenant_id=ctx["tenant_id"], inquiry_id=result.inserted_id, metadata={"comparison_group_id": group_id})
        created.append({"inquiry_id": str(result.inserted_id), "vendor_id": str(vendor_id)})
    return {
        "status": "success", "message": f"RFQ sent to {len(created)} vendor(s).",
        "comparison_group_id": group_id, "created": created,
    }


# MULTI-VENDOR COMPARISON — buyer sends the same spec to several vendors at
# once, sees all their quotes side by side, converts whichever one wins.
# ═══════════════════════════════════════════════════════════════════════════

@router.post("/inquiries/compare", status_code=201)
async def create_comparison_inquiries(payload: dict, ctx: dict = Depends(get_hq_tenant)):
    """
    Buyer picks several DIFFERENT catalogue items (one per vendor — each
    vendor has their own catalogue entry for "casual t-shirt", they're
    separate documents) and sends the same requested size/color/qty/price
    to all of them in one action.

    This creates N independent inquiry documents (reusing the exact same
    single-inquiry logic and validation as create_inquiry above — nothing
    about the negotiation/PO-safety gate changes), all tagged with a shared
    comparison_group_id so the buyer can pull them back up side by side
    once vendors start responding.

    Body:
        {
          "catalogue_item_ids": ["...", "...", "..."],
          "requested_size": "M", "requested_color": "Navy",
          "requested_qty": 200, "requested_price": 250,
          "buyer_note": "Need by end of month"
        }
    """
    request_items = payload.get("items") or []
    if request_items:
        normalized_items = [entry for entry in request_items if isinstance(entry, dict)]
    else:
        # Backward compatibility for clients using shared fields.
        normalized_items = [
            {"catalogue_item_id": item_id, **payload}
            for item_id in (payload.get("catalogue_item_ids") or [])
        ]
    if len(normalized_items) < 2:
        raise HTTPException(status_code=400, detail="At least 2 catalogue items are required for a comparison; use POST /inquiries for a single product.")

    group_id = str(uuid.uuid4())
    created  = []
    skipped  = []

    for request_item in normalized_items:
        item_id = request_item.get("catalogue_item_id")
        if not ObjectId.is_valid(item_id):
            skipped.append({"catalogue_item_id": item_id, "reason": "Invalid ID"})
            continue

        item = await vendor_catalogue_collection.find_one({"_id": ObjectId(item_id)})
        if not item:
            skipped.append({"catalogue_item_id": item_id, "reason": "Catalogue item not found"})
            continue

        # Same ownership check as the single-vendor create_inquiry route —
        # a buyer can't fan out to a vendor they don't have an approved
        # relationship with just because they know a catalogue_item_id.
        approved_link = await vendor_tenant_links_collection.find_one({
            "vendor_id": item["vendor_id"],
            "tenant_id": ctx["tenant_id"],
            "status":    "Approved",
        })
        if not approved_link:
            skipped.append({"catalogue_item_id": item_id, "reason": "No approved relationship with this vendor"})
            continue

        doc = {
            "catalogue_item_id":  item["_id"],
            "vendor_id":          item["vendor_id"],
            "tenant_id":          ctx["tenant_id"],
            "item_name":          item.get("item_name", ""),
            "item_image":         (item.get("images") or [None])[0],
            "requested_size":     (request_item.get("requested_size") or "").strip(),
            "requested_color":    (request_item.get("requested_color") or "").strip(),
            "requested_qty":      max(0, int(request_item.get("requested_qty") or 0)),
            "requested_price":    max(0.0, float(request_item.get("requested_price") or 0)),
            "buyer_note":         (request_item.get("buyer_note") or "").strip(),
            "response_deadline":  (request_item.get("response_deadline") or payload.get("response_deadline") or "").strip(),
            "status":             "Pending",
            "comparison_group_id": group_id,   # ← the only thing that differs from a single inquiry
            "created_by":         ctx.get("admin_id"),
            "created_at":         datetime.utcnow(),
            "vendor_response":    None,
        }
        result = await catalogue_inquiries_collection.insert_one(doc)
        await notify_vendor(item["vendor_id"], "rfq_received", "New quote request", f"A buyer requested a quote for {item.get('item_name', 'a catalogue item')}.", tenant_id=ctx["tenant_id"], inquiry_id=result.inserted_id, metadata={"comparison_group_id": group_id})
        created.append({"inquiry_id": str(result.inserted_id), "vendor_id": str(item["vendor_id"]), "item_name": item.get("item_name", "")})

    if not created:
        raise HTTPException(status_code=400, detail="None of the given catalogue items could be sent an inquiry.", headers={"X-Skipped": str(len(skipped))})

    return {
        "status":             "success",
        "message":            f"Sent to {len(created)} vendor(s)." + (f" {len(skipped)} skipped." if skipped else ""),
        "comparison_group_id": group_id,
        "created":            created,
        "skipped":            skipped,
    }


@router.get("/inquiries/compare/{group_id}")
async def get_comparison_group(group_id: str, ctx: dict = Depends(get_hq_tenant)):
    """
    Returns every inquiry in a comparison group, tenant-scoped, so the
    buyer can see all vendors' quotes for the same request side by side.
    Ordered by confirmed price ascending once responded (cheapest first),
    Pending ones last since there's nothing to compare yet.
    """
    rows = []
    async for doc in catalogue_inquiries_collection.find({
        "comparison_group_id": group_id,
        "tenant_id": ctx["tenant_id"],
    }):
        row = _serialize_inquiry(doc)
        vendor = await vendors_collection.find_one({"_id": doc["vendor_id"]})
        row["vendor_name"] = (vendor or {}).get("name") or (vendor or {}).get("vendor_name", "Unknown Vendor")
        rows.append(row)

    if not rows:
        raise HTTPException(status_code=404, detail="No inquiries found for this comparison group.")

    # ⚠️ FIXED: previously sorted by price alone, globally. If a buyer
    # selected multiple DIFFERENT products in one compare batch (e.g. 3
    # vendors' "Cotton T-Shirt" quotes AND 3 vendors' "Floral Kurti"
    # quotes, all sent as one comparison_group_id), those got interleaved
    # by price with zero regard for which product each quote was even
    # for — a cheap kurti quote could sort above an expensive t-shirt
    # quote, making "comparison" meaningless since you can't compare a
    # t-shirt's price against a kurti's. Now groups by item_name first
    # (alphabetically, for stable ordering), THEN sorts by price within
    # each product group — so all Cotton T-Shirt quotes stay together,
    # all Floral Kurti quotes stay together, each internally cheapest-first.
    def sort_key(r):
        vr = r.get("vendor_response") or {}
        has_price = r["status"] == "Responded" and vr.get("confirmed_price")
        return (
            r.get("item_name", ""),
            0 if has_price else 1,
            vr.get("confirmed_price", float("inf")) if has_price else 0,
        )

    rows.sort(key=sort_key)

    return {"status": "success", "comparison_group_id": group_id, "count": len(rows), "data": rows}
